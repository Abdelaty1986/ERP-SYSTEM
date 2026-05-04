import shutil
import sys
import tempfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
TMP_ROOT = PROJECT_ROOT / "tests" / "tmp"
TMP_ROOT.mkdir(parents=True, exist_ok=True)
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import app as appmod


def assert_true(condition, message):
    if not condition:
        raise AssertionError(message)


def fetchone(cur, sql, params=()):
    cur.execute(sql, params)
    return cur.fetchone()


def workbook_rows(response):
    wb = load_workbook(BytesIO(response.data))
    sheet = wb.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        values = [str(value).strip() for value in row if value not in (None, "")]
        if values:
            rows.append(" | ".join(values))
    return rows


def main():
    source_db = PROJECT_ROOT / "database.db"
    temp_dir = Path(tempfile.mkdtemp(prefix="erp-integration-", dir=str(TMP_ROOT)))
    temp_db = temp_dir / "database_test.db"
    shutil.copy2(source_db, temp_db)

    old_db_path = appmod.DB_PATH
    old_module_db_path = appmod.MODULE_DEPS.get("DB_PATH")

    appmod.DB_PATH = str(temp_db)
    appmod.MODULE_DEPS["DB_PATH"] = str(temp_db)
    appmod.run_migrations(str(temp_db))
    appmod.init_db()
    appmod.app.config["TESTING"] = True

    client = appmod.app.test_client()
    with client.session_transaction() as session:
        session["user_id"] = 1
        session["username"] = "codex-invoice-test"
        session["role"] = "admin"

    conn = appmod.db()
    cur = conn.cursor()

    cur.execute("INSERT INTO customers(name,tax_id) VALUES (?,?)", ("عميل اختبار ضرائب السطور", "C-TAX-001"))
    customer_id = cur.lastrowid
    cur.execute("INSERT INTO suppliers(name,tax_id) VALUES (?,?)", ("مورد اختبار ضرائب السطور", "S-TAX-001"))
    supplier_id = cur.lastrowid

    products = [
        ("VAT-ONLY", "صنف خاضع VAT فقط", 10, 20),
        ("NO-TAX", "صنف غير خاضع", 8, 15),
        ("VAT-WHT", "صنف VAT وخصم", 5, 30),
        ("PUR-VAT", "شراء VAT فقط", 12, 22),
        ("PUR-NO-TAX", "شراء بدون ضريبة", 9, 16),
        ("PUR-VAT-WHT", "شراء VAT وخصم", 7, 18),
    ]
    product_ids = {}
    for code, name, purchase_price, sale_price in products:
        cur.execute(
            "INSERT INTO products(code,name,unit,purchase_price,sale_price,stock_quantity) VALUES (?,?,?,?,?,?)",
            (code, name, "علبة", purchase_price, sale_price, 120),
        )
        product_ids[code] = cur.lastrowid
    conn.commit()

    sales_page = client.get("/sales")
    assert_true(sales_page.status_code == 200, "sales page failed")
    sales_html = sales_page.get_data(as_text=True)
    assert_true('id="sales-summary-net"' in sales_html, "sales summary panel missing")
    assert_true("initInvoiceWorkspace" in sales_html, "sales live totals JS missing")
    assert_true("line-subtotal" in sales_html and "line-net" in sales_html, "sales line totals are not rendered in HTML")

    purchases_page = client.get("/purchases")
    assert_true(purchases_page.status_code == 200, "purchases page failed")
    purchases_html = purchases_page.get_data(as_text=True)
    assert_true('id="purchase-summary-net"' in purchases_html, "purchase summary panel missing")
    assert_true("initInvoiceWorkspace" in purchases_html, "purchase live totals JS missing")

    response = client.post(
        "/sales",
        data={
            "date": "2026-05-01",
            "invoice_number": "INV-TAX-001",
            "customer_id": str(customer_id),
            "payment_type": "credit",
            "product_id[]": [str(product_ids["VAT-ONLY"]), str(product_ids["NO-TAX"]), str(product_ids["VAT-WHT"])],
            "unit_id[]": ["", "", ""],
            "quantity[]": ["2", "3", "1"],
            "unit_price[]": ["20", "15", "30"],
            "vat_applicable[]": ["1", "0", "1"],
            "vat_rate[]": ["14", "0", "14"],
            "withholding_applicable[]": ["0", "0", "1"],
            "withholding_rate[]": ["0", "0", "5"],
        },
        follow_redirects=False,
    )
    assert_true(response.status_code in (302, 303), "sales invoice creation failed")

    sales_invoice_id = fetchone(cur, "SELECT id FROM sales_invoices WHERE invoice_number='INV-TAX-001'")[0]
    sales_header = fetchone(
        cur,
        """
        SELECT subtotal, COALESCE(vat_total, tax_amount), COALESCE(withholding_total, withholding_amount), COALESCE(net_total, grand_total)
        FROM sales_invoices
        WHERE id=?
        """,
        (sales_invoice_id,),
    )
    assert_true(tuple(round(float(value or 0), 2) for value in sales_header) == (115.0, 9.8, 1.5, 123.3), "sales header totals are incorrect")
    sales_lines = cur.execute(
        """
        SELECT product_id, subtotal, vat_applicable, vat_amount, withholding_applicable, withholding_amount, line_net, selected_unit, conversion_factor, base_qty
        FROM sales_invoice_lines
        WHERE invoice_id=?
        ORDER BY id
        """,
        (sales_invoice_id,),
    ).fetchall()
    assert_true(len(sales_lines) == 3, "sales line count mismatch")
    assert_true(round(float(sales_lines[0][1] or 0), 2) == 40.0 and round(float(sales_lines[0][3] or 0), 2) == 5.6, "VAT-only line values are incorrect")
    assert_true(int(sales_lines[1][2] or 0) == 0 and round(float(sales_lines[1][3] or 0), 2) == 0.0, "non-taxable line should not include VAT")
    assert_true(int(sales_lines[2][4] or 0) == 1 and round(float(sales_lines[2][5] or 0), 2) == 1.5 and round(float(sales_lines[2][6] or 0), 2) == 32.7, "mixed line values are incorrect")
    assert_true(all(line[7] for line in sales_lines), "selected unit was not stored on sales lines")
    assert_true(all(float(line[8] or 0) >= 1 for line in sales_lines), "conversion factor was not stored on sales lines")
    assert_true(all(float(line[9] or 0) > 0 for line in sales_lines), "base quantity was not preserved on sales lines")

    duplicate_response = client.post(
        "/sales",
        data={
            "date": "2026-05-01",
            "invoice_number": "INV-TAX-001",
            "customer_id": str(customer_id),
            "payment_type": "cash",
            "product_id[]": [str(product_ids["VAT-ONLY"])],
            "unit_id[]": [""],
            "quantity[]": ["1"],
            "unit_price[]": ["20"],
            "vat_applicable[]": ["1"],
            "vat_rate[]": ["14"],
            "withholding_applicable[]": ["0"],
            "withholding_rate[]": ["0"],
        },
        follow_redirects=True,
    )
    assert_true(duplicate_response.status_code == 200, "duplicate sales invoice request failed")
    assert_true(fetchone(cur, "SELECT COUNT(*) FROM sales_invoices WHERE invoice_number='INV-TAX-001'")[0] == 1, "duplicate sales invoice was not blocked")

    purchase_response = client.post(
        "/purchases",
        data={
            "date": "2026-05-02",
            "invoice_number": "PINV-TAX-001",
            "supplier_invoice_no": "SUP-INV-001",
            "supplier_invoice_date": "2026-05-02",
            "supplier_id": str(supplier_id),
            "payment_type": "credit",
            "product_id[]": [str(product_ids["PUR-VAT"]), str(product_ids["PUR-NO-TAX"]), str(product_ids["PUR-VAT-WHT"])],
            "unit_id[]": ["", "", ""],
            "quantity[]": ["2", "3", "1"],
            "unit_price[]": ["12", "9", "7"],
            "vat_applicable[]": ["1", "0", "1"],
            "vat_rate[]": ["14", "0", "14"],
            "withholding_applicable[]": ["0", "0", "1"],
            "withholding_rate[]": ["0", "0", "2"],
        },
        follow_redirects=False,
    )
    assert_true(purchase_response.status_code in (302, 303), "purchase invoice creation failed")

    purchase_invoice_id = fetchone(cur, "SELECT id FROM purchase_invoices WHERE invoice_number='PINV-TAX-001'")[0]
    purchase_header = fetchone(
        cur,
        """
        SELECT subtotal, COALESCE(vat_total, tax_amount), COALESCE(withholding_total, withholding_amount), COALESCE(net_total, grand_total)
        FROM purchase_invoices
        WHERE id=?
        """,
        (purchase_invoice_id,),
    )
    assert_true(tuple(round(float(value or 0), 2) for value in purchase_header) == (58.0, 4.34, 0.14, 62.2), "purchase header totals are incorrect")
    purchase_lines = cur.execute(
        """
        SELECT product_id, subtotal, vat_applicable, vat_amount, withholding_applicable, withholding_amount, line_net
        FROM purchase_invoice_lines
        WHERE invoice_id=?
        ORDER BY id
        """,
        (purchase_invoice_id,),
    ).fetchall()
    assert_true(len(purchase_lines) == 3, "purchase line count mismatch")
    assert_true(int(purchase_lines[1][2] or 0) == 0 and round(float(purchase_lines[1][3] or 0), 2) == 0.0, "purchase non-taxable line should not include VAT")

    journal_balance = fetchone(
        cur,
        """
        SELECT ROUND(COALESCE(SUM(debit), 0), 2), ROUND(COALESCE(SUM(credit), 0), 2)
        FROM ledger
        WHERE journal_id IN (SELECT id FROM journal WHERE source_type='sales' AND source_id=?)
        """,
        (sales_invoice_id,),
    )
    assert_true(journal_balance[0] == journal_balance[1] and journal_balance[0] > 0, "sales journal is not balanced")

    vat_report = client.get("/reports/vat?format=excel")
    assert_true(vat_report.status_code == 200, "VAT report export failed")
    vat_rows = workbook_rows(vat_report)
    assert_true(any("صنف خاضع VAT فقط" in row for row in vat_rows), "VAT report is missing taxable sales line")
    assert_true(any("صنف VAT وخصم" in row for row in vat_rows), "VAT report is missing mixed sales line")
    assert_true(not any("صنف غير خاضع" in row for row in vat_rows), "VAT report included a non-taxable line")

    withholding_report = client.get("/reports/withholding-tax?format=excel")
    assert_true(withholding_report.status_code == 200, "withholding report export failed")
    withholding_rows = workbook_rows(withholding_report)
    assert_true(any("صنف VAT وخصم" in row for row in withholding_rows), "withholding report is missing mixed line")
    assert_true(not any("صنف خاضع VAT فقط" in row for row in withholding_rows), "withholding report included a line without withholding")

    conn.close()
    appmod.DB_PATH = old_db_path
    appmod.MODULE_DEPS["DB_PATH"] = old_module_db_path
    shutil.rmtree(temp_dir, ignore_errors=True)
    print("invoice_workflow_test: ok")


if __name__ == "__main__":
    main()
