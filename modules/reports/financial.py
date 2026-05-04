from flask import render_template_string, request
from io import BytesIO

from flask import flash, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook
from openpyxl.styles import Font


def _account_balances_by_type(cur, account_type):
    cur.execute(
        """
        SELECT a.code,a.name,
               COALESCE(SUM(l.debit),0) AS debit,
               COALESCE(SUM(l.credit),0) AS credit
        FROM accounts a
        LEFT JOIN ledger l ON l.account_id=a.id
        WHERE a.type=?
        GROUP BY a.id
        ORDER BY a.code
        """,
        (account_type,),
    )
    return cur.fetchall()


def _debit_balance_rows(rows):
    result = []
    for code, name, debit, credit in rows:
        amount = (debit or 0) - (credit or 0)
        if abs(amount) > 0.0001:
            result.append((code, name, amount))
    return result


def _credit_balance_rows(rows):
    result = []
    for code, name, debit, credit in rows:
        amount = (credit or 0) - (debit or 0)
        if abs(amount) > 0.0001:
            result.append((code, name, amount))
    return result


def _tax_report_excel(filename, title, rows, total_label, total_value):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.sheet_view.rightToLeft = True
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    headers = ["ط§ظ„طھط§ط±ظٹط®", "ط±ظ‚ظ… ط§ظ„ظپط§طھظˆط±ط©", "ط§ظ„ط¬ظ‡ط©", "ط§ظ„طµظ†ظپ", "ط§ظ„ظˆط­ط¯ط©", "ظ‚ظٹظ…ط© ط§ظ„طµظ†ظپ", "ط§ظ„ظ†ط³ط¨ط©", "ط§ظ„ط¶ط±ظٹط¨ط©", "ط§ظ„ظ†ظˆط¹", "ط§ظ„ط¥ط¬ظ…ط§ظ„ظٹ"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=idx, value=header).font = Font(bold=True)
    row_idx = 3
    for row in rows:
        row_idx += 1
        for idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=idx, value=value)
    ws.cell(row=row_idx + 2, column=1, value=total_label).font = Font(bold=True)
    ws.cell(row=row_idx + 2, column=2, value=total_value)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def build_balance_sheet_report_view(deps):
    db = deps["db"]

    def balance_sheet_report():
        conn = db()
        cur = conn.cursor()
        assets = _debit_balance_rows(_account_balances_by_type(cur, "ط£طµظˆظ„"))
        liabilities = _credit_balance_rows(_account_balances_by_type(cur, "ط®طµظˆظ…"))
        equity = _credit_balance_rows(_account_balances_by_type(cur, "ط­ظ‚ظˆظ‚ ظ…ظ„ظƒظٹط©"))
        revenue_total = sum(row[2] for row in _credit_balance_rows(_account_balances_by_type(cur, "ط¥ظٹط±ط§ط¯ط§طھ")))
        expense_total = sum(row[2] for row in _debit_balance_rows(_account_balances_by_type(cur, "ظ…طµط±ظˆظپط§طھ")))
        net_income = revenue_total - expense_total
        if abs(net_income) > 0.0001:
            equity.append(("3400", "طµط§ظپظٹ ط±ط¨ط­ ط£ظˆ ط®ط³ط§ط±ط© ط§ظ„ظپطھط±ط©", net_income))
        total_assets = sum(row[2] for row in assets)
        total_liabilities = sum(row[2] for row in liabilities)
        total_equity = sum(row[2] for row in equity)
        conn.close()
        return render_template(
            "balance_sheet.html",
            assets=assets,
            liabilities=liabilities,
            equity=equity,
            total_assets=total_assets,
            total_liabilities=total_liabilities,
            total_equity=total_equity,
            difference=total_assets - (total_liabilities + total_equity),
        )

    return balance_sheet_report


def build_cash_flow_report_view(deps):
    db = deps["db"]

    def cash_flow_report():
        conn = db()
        cur = conn.cursor()
        cash_codes = ("1100", "1110", "1120", "1200", "1210")
        placeholders = ",".join(["?"] * len(cash_codes))
        cur.execute(
            f"""
            SELECT j.date,j.description,da.code,da.name,ca.code,ca.name,j.amount
            FROM journal j
            JOIN accounts da ON da.id=j.debit_account_id
            JOIN accounts ca ON ca.id=j.credit_account_id
            WHERE j.status='posted'
              AND (da.code IN ({placeholders}) OR ca.code IN ({placeholders}))
            ORDER BY j.date,j.id
            """,
            (*cash_codes, *cash_codes),
        )
        rows = []
        totals = {"operating": 0, "investing": 0, "financing": 0}
        for date_value, desc, debit_code, debit_name, credit_code, credit_name, amount in cur.fetchall():
            if debit_code in cash_codes:
                direction = "in"
                other_code = credit_code
                other_name = credit_name
                signed_amount = amount
            else:
                direction = "out"
                other_code = debit_code
                other_name = debit_name
                signed_amount = -amount

            if other_code.startswith("18") or other_code.startswith("145"):
                activity = "investing"
                activity_label = "ط§ط³طھط«ظ…ط§ط±ظٹط©"
            elif other_code.startswith("24") or other_code.startswith("25") or other_code.startswith("3"):
                activity = "financing"
                activity_label = "طھظ…ظˆظٹظ„ظٹط©"
            else:
                activity = "operating"
                activity_label = "طھط´ط؛ظٹظ„ظٹط©"
            totals[activity] += signed_amount
            rows.append((date_value, desc, activity_label, other_code, other_name, direction, amount, signed_amount))

        cur.execute(
            f"""
            SELECT COALESCE(SUM(l.debit),0)-COALESCE(SUM(l.credit),0)
            FROM ledger l
            JOIN accounts a ON a.id=l.account_id
            WHERE a.code IN ({placeholders})
            """,
            cash_codes,
        )
        ending_cash = cur.fetchone()[0]
        net_cash_flow = sum(totals.values())
        beginning_cash = ending_cash - net_cash_flow
        conn.close()
        return render_template(
            "cash_flow.html",
            rows=rows,
            totals=totals,
            beginning_cash=beginning_cash,
            net_cash_flow=net_cash_flow,
            ending_cash=ending_cash,
        )

    return cash_flow_report


def build_cost_center_report_view(deps):
    db = deps["db"]

    def cost_center_report():
        conn = db()
        cur = conn.cursor()
        cur.execute(
            """
            WITH lines AS (
                SELECT cc.code AS center_code,cc.name AS center_name,a.type AS account_type,
                       j.amount AS debit,0 AS credit
                FROM journal j
                JOIN cost_centers cc ON cc.id=j.cost_center_id
                JOIN accounts a ON a.id=j.debit_account_id
                WHERE j.status='posted'
                UNION ALL
                SELECT cc.code AS center_code,cc.name AS center_name,a.type AS account_type,
                       0 AS debit,j.amount AS credit
                FROM journal j
                JOIN cost_centers cc ON cc.id=j.cost_center_id
                JOIN accounts a ON a.id=j.credit_account_id
                WHERE j.status='posted'
            )
            SELECT center_code,center_name,
                   SUM(CASE WHEN account_type='ط¥ظٹط±ط§ط¯ط§طھ' THEN credit-debit ELSE 0 END) AS revenue,
                   SUM(CASE WHEN account_type='ظ…طµط±ظˆظپط§طھ' THEN debit-credit ELSE 0 END) AS expense
            FROM lines
            GROUP BY center_code,center_name
            ORDER BY center_code,center_name
            """
        )
        rows = cur.fetchall()
        totals = {"revenue": sum(row[2] or 0 for row in rows), "expense": sum(row[3] or 0 for row in rows)}
        conn.close()
        return render_template("cost_center_report.html", rows=rows, totals=totals)

    return cost_center_report


def build_opening_balances_view(deps):
    db = deps["db"]
    parse_positive_amount = deps["parse_positive_amount"]
    ensure_open_period = deps["ensure_open_period"]
    get_account_id = deps["get_account_id"]
    log_action = deps["log_action"]
    rebuild_ledger = deps["rebuild_ledger"]

    def opening_balances():
        conn = db()
        cur = conn.cursor()
        if request.method == "POST":
            date_value = request.form.get("date", "").strip()
            account_id = request.form.get("account_id")
            side = request.form.get("side", "debit")
            amount = parse_positive_amount(request.form.get("amount"))
            notes = request.form.get("notes", "").strip()
            if not date_value or not account_id or amount <= 0 or side not in ("debit", "credit"):
                flash("ط±ط§ط¬ط¹ ط¨ظٹط§ظ†ط§طھ ط§ظ„ط±طµظٹط¯ ط§ظ„ط§ظپطھطھط§ط­ظٹ.", "danger")
            else:
                try:
                    ensure_open_period(cur, date_value)
                except ValueError as exc:
                    flash(str(exc), "danger")
                    conn.close()
                    return redirect(url_for("opening_balances"))
                opening_id = get_account_id(cur, "1900")
                if side == "debit":
                    debit_id, credit_id = account_id, opening_id
                else:
                    debit_id, credit_id = opening_id, account_id
                cur.execute(
                    """
                    INSERT INTO journal(date,description,debit_account_id,credit_account_id,amount,status,source_type)
                    VALUES (?,?,?,?,?,'posted','opening')
                    """,
                    (date_value, f"ط±طµظٹط¯ ط§ظپطھطھط§ط­ظٹ - {notes}", debit_id, credit_id, amount),
                )
                journal_id = cur.lastrowid
                log_action(cur, "create", "opening_balance", journal_id, f"amount={amount}")
                conn.commit()
                conn.close()
                rebuild_ledger()
                flash("طھظ… طھط³ط¬ظٹظ„ ط§ظ„ط±طµظٹط¯ ط§ظ„ط§ظپطھطھط§ط­ظٹ.", "success")
                return redirect(url_for("opening_balances"))
        cur.execute("SELECT id,code,name,type FROM accounts ORDER BY code")
        accounts_rows = cur.fetchall()
        cur.execute(
            """
            SELECT j.date,j.description,a1.name,a2.name,j.amount
            FROM journal j
            JOIN accounts a1 ON a1.id=j.debit_account_id
            JOIN accounts a2 ON a2.id=j.credit_account_id
            WHERE j.source_type='opening'
            ORDER BY j.id DESC
            """
        )
        rows = cur.fetchall()
        conn.close()
        return render_template("opening_balances.html", accounts=accounts_rows, rows=rows)

    return opening_balances


def build_year_end_view(deps):
    db = deps["db"]
    ensure_open_period = deps["ensure_open_period"]
    create_auto_journal = deps["create_auto_journal"]
    mark_journal_source = deps["mark_journal_source"]
    log_action = deps["log_action"]
    rebuild_ledger = deps["rebuild_ledger"]

    def year_end():
        conn = db()
        cur = conn.cursor()
        if request.method == "POST":
            fiscal_year = request.form.get("fiscal_year", "").strip()
            closing_date = request.form.get("closing_date", "").strip()
            notes = request.form.get("notes", "").strip()
            if not fiscal_year or not closing_date:
                flash("ط§ظ„ط³ظ†ط© ظˆطھط§ط±ظٹط® ط§ظ„ط¥ظ‚ظپط§ظ„ ظ…ط·ظ„ظˆط¨ط§ظ†.", "danger")
            else:
                cur.execute("SELECT id FROM year_end_closings WHERE fiscal_year=?", (fiscal_year,))
                if cur.fetchone():
                    flash("طھظ… ط¥ظ‚ظپط§ظ„ ظ‡ط°ظ‡ ط§ظ„ط³ظ†ط© ظ…ظ† ظ‚ط¨ظ„.", "danger")
                else:
                    try:
                        ensure_open_period(cur, closing_date)
                    except ValueError as exc:
                        flash(str(exc), "danger")
                        conn.close()
                        return redirect(url_for("year_end"))
                    revenue_rows = _account_balances_by_type(cur, "ط¥ظٹط±ط§ط¯ط§طھ")
                    expense_rows = _account_balances_by_type(cur, "ظ…طµط±ظˆظپط§طھ")
                    revenue_total = 0
                    expense_total = 0
                    closing_journal_ids = []

                    for code, name, debit, credit in revenue_rows:
                        amount = (credit or 0) - (debit or 0)
                        revenue_total += amount
                        if abs(amount) <= 0.0001:
                            continue
                        if amount > 0:
                            closing_journal_ids.append(create_auto_journal(cur, closing_date, f"ط¥ظ‚ظپط§ظ„ ط¥ظٹط±ط§ط¯ {name} - {fiscal_year}", code, "3400", amount))
                        else:
                            closing_journal_ids.append(create_auto_journal(cur, closing_date, f"ط¥ظ‚ظپط§ظ„ ظ…ط±ط¯ظˆط¯/ط®طµظ… {name} - {fiscal_year}", "3400", code, abs(amount)))

                    for code, name, debit, credit in expense_rows:
                        amount = (debit or 0) - (credit or 0)
                        expense_total += amount
                        if abs(amount) <= 0.0001:
                            continue
                        if amount > 0:
                            closing_journal_ids.append(create_auto_journal(cur, closing_date, f"ط¥ظ‚ظپط§ظ„ ظ…طµط±ظˆظپ {name} - {fiscal_year}", "3400", code, amount))
                        else:
                            closing_journal_ids.append(create_auto_journal(cur, closing_date, f"ط¥ظ‚ظپط§ظ„ ط¹ظƒط³ ظ…طµط±ظˆظپ {name} - {fiscal_year}", code, "3400", abs(amount)))

                    net_income = revenue_total - expense_total
                    if abs(net_income) < 0.0001:
                        journal_id = None
                    elif net_income > 0:
                        journal_id = create_auto_journal(cur, closing_date, f"طھط±ط­ظٹظ„ طµط§ظپظٹ ط±ط¨ط­ {fiscal_year}", "3400", "3500", net_income)
                        closing_journal_ids.append(journal_id)
                    else:
                        journal_id = create_auto_journal(cur, closing_date, f"طھط±ط­ظٹظ„ طµط§ظپظٹ ط®ط³ط§ط±ط© {fiscal_year}", "3500", "3400", abs(net_income))
                        closing_journal_ids.append(journal_id)
                    cur.execute(
                        """
                        INSERT INTO year_end_closings(fiscal_year,closing_date,revenue_total,expense_total,net_income,journal_id,notes)
                        VALUES (?,?,?,?,?,?,?)
                        """,
                        (fiscal_year, closing_date, revenue_total, expense_total, net_income, journal_id, notes),
                    )
                    closing_id = cur.lastrowid
                    mark_journal_source(cur, "year_end", closing_id, *closing_journal_ids)
                    log_action(cur, "close", "year_end", closing_id, fiscal_year)
                    conn.commit()
                    conn.close()
                    rebuild_ledger()
                    flash("طھظ… طھظ†ظپظٹط° ظ‚ظٹظˆط¯ ط¥ظ‚ظپط§ظ„ ط§ظ„ط³ظ†ط© ظˆطھط±ط­ظٹظ„ ط§ظ„ظ†طھظٹط¬ط©.", "success")
                    return redirect(url_for("year_end"))
        cur.execute("SELECT id,fiscal_year,closing_date,revenue_total,expense_total,net_income,status FROM year_end_closings ORDER BY fiscal_year DESC")
        rows = cur.fetchall()
        conn.close()
        return render_template("year_end.html", rows=rows)

    return year_end


def build_profit_loss_report_view(deps):
    db = deps["db"]

    def profit_loss_report():
        conn = db()
        cur = conn.cursor()
        cur.execute("SELECT COALESCE(SUM(total),0), COALESCE(SUM(cost_total),0) FROM sales_invoices WHERE status='posted'")
        sales_total, cost_total = cur.fetchone()
        cur.execute("SELECT COALESCE(SUM(amount),0) FROM financial_sales_invoices WHERE status='posted'")
        sales_total += cur.fetchone()[0]
        cur.execute("SELECT COALESCE(SUM(total),0) FROM purchase_invoices WHERE status='posted'")
        purchases_total = cur.fetchone()[0]
        cur.execute(
            """
            SELECT a.name, COALESCE(SUM(l.debit),0) - COALESCE(SUM(l.credit),0) AS amount
            FROM accounts a
            LEFT JOIN ledger l ON a.id=l.account_id
            WHERE a.type='ظ…طµط±ظˆظپط§طھ' AND a.code <> '6100'
            GROUP BY a.id
            ORDER BY a.code
            """
        )
        expense_rows = cur.fetchall()
        expenses_total = sum(row[1] for row in expense_rows)
        gross_profit = sales_total - cost_total
        net_profit = gross_profit - expenses_total
        conn.close()
        return render_template(
            "profit_loss.html",
            sales_total=sales_total,
            cost_total=cost_total,
            purchases_total=purchases_total,
            gross_profit=gross_profit,
            expense_rows=expense_rows,
            expenses_total=expenses_total,
            net_profit=net_profit,
        )

    return profit_loss_report

def _table_columns(cur, table_name):
    try:
        return {row[1] for row in cur.execute(f"PRAGMA table_info({table_name})").fetchall()}
    except Exception:
        return set()


def _column_value_expr(cur, table_name, alias, column_names, default="0", nullif_empty=False):
    columns = _table_columns(cur, table_name)
    expressions = []
    for column_name in column_names:
        if column_name in columns:
            if nullif_empty:
                expressions.append(f"NULLIF({alias}.{column_name}, '')")
            else:
                expressions.append(f"{alias}.{column_name}")
    if expressions:
        return f"COALESCE({', '.join(expressions)}, {default})"
    return default


def _report_tax_number_expr(cur, table_name, alias, fallback="'ط؛ظٹط± ظ…ط³ط¬ظ„'"):
    return _column_value_expr(
        cur,
        table_name,
        alias,
        ("tax_registration_number", "tax_number", "tax_id", "vat_number"),
        fallback,
        nullif_empty=True,
    )


def _financial_report_excel(filename, headers, rows, title):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.sheet_view.rightToLeft = True
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=idx, value=header).font = Font(bold=True)
    for row_idx, row in enumerate(rows, start=4):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def build_vat_report_view(deps):
    db = deps["db"]

    def vat_report():
        conn = db()
        cur = conn.cursor()
        date_from = (request.args.get("date_from") or "").strip()
        date_to = (request.args.get("date_to") or "").strip()

        sales_tax_expr = _report_tax_number_expr(cur, "customers", "c", "'ط؛ظٹط± ظ…ط³ط¬ظ„'")
        supplier_tax_expr = _report_tax_number_expr(cur, "suppliers", "sup", "'ط؛ظٹط± ظ…ط³ط¬ظ„'")

        sales_qty_expr = _column_value_expr(cur, "sales_invoice_lines", "l", ("qty", "quantity"), "0")
        sales_price_expr = _column_value_expr(cur, "sales_invoice_lines", "l", ("unit_price", "price"), "0")
        sales_unit_expr = _column_value_expr(cur, "sales_invoice_lines", "l", ("selected_unit", "unit_name"), "p.unit", nullif_empty=True)
        sales_line_total_expr = _column_value_expr(cur, "sales_invoice_lines", "l", ("subtotal", "total", "line_total"), f"(({sales_qty_expr}) * ({sales_price_expr}))")
        sales_vat_applicable_expr = _column_value_expr(cur, "sales_invoice_lines", "l", ("vat_applicable", "vat_enabled"), "1")
        sales_vat_amount_expr = _column_value_expr(cur, "sales_invoice_lines", "l", ("vat_amount", "tax_amount"), "0")

        purchase_qty_expr = _column_value_expr(cur, "purchase_invoice_lines", "l", ("qty", "quantity"), "0")
        purchase_price_expr = _column_value_expr(cur, "purchase_invoice_lines", "l", ("unit_price", "price"), "0")
        purchase_unit_expr = _column_value_expr(cur, "purchase_invoice_lines", "l", ("selected_unit", "unit_name"), "pr.unit", nullif_empty=True)
        purchase_line_total_expr = _column_value_expr(cur, "purchase_invoice_lines", "l", ("subtotal", "total", "line_total"), f"(({purchase_qty_expr}) * ({purchase_price_expr}))")
        purchase_vat_applicable_expr = _column_value_expr(cur, "purchase_invoice_lines", "l", ("vat_applicable", "vat_enabled"), "1")
        purchase_vat_amount_expr = _column_value_expr(cur, "purchase_invoice_lines", "l", ("vat_amount", "tax_amount"), "0")

        sales_invoice_no_expr = _column_value_expr(cur, "sales_invoices", "s", ("invoice_number", "doc_no"), "'SI-' || printf('%06d', s.id)", nullif_empty=True)
        purchase_invoice_no_expr = _column_value_expr(cur, "purchase_invoices", "p", ("invoice_number", "doc_no", "supplier_invoice_no"), "'PI-' || printf('%06d', p.id)", nullif_empty=True)

        sales_conditions = ["s.status='posted'", f"COALESCE({sales_vat_applicable_expr}, 1) = 1", f"COALESCE({sales_vat_amount_expr}, 0) > 0"]
        purchase_conditions = ["p.status='posted'", f"COALESCE({purchase_vat_applicable_expr}, 1) = 1", f"COALESCE({purchase_vat_amount_expr}, 0) > 0"]
        params_sales = []
        params_purchases = []
        if date_from:
            sales_conditions.append("s.date >= ?")
            purchase_conditions.append("p.date >= ?")
            params_sales.append(date_from)
            params_purchases.append(date_from)
        if date_to:
            sales_conditions.append("s.date <= ?")
            purchase_conditions.append("p.date <= ?")
            params_sales.append(date_to)
            params_purchases.append(date_to)

        sales_where = " AND ".join(sales_conditions)
        purchases_where = " AND ".join(purchase_conditions)

        cur.execute(
            f"""
            SELECT
                s.date,
                'ظ…ط¨ظٹط¹ط§طھ' AS doc_type,
                {sales_invoice_no_expr} AS doc_no,
                COALESCE(c.name, 'ط¹ظ…ظٹظ„ ظ†ظ‚ط¯ظٹ') AS party_name,
                {sales_tax_expr} AS tax_number,
                COALESCE(p.name, '') AS item_name,
                {sales_qty_expr} AS qty,
                {sales_price_expr} AS unit_price,
                COALESCE({sales_unit_expr}, 'ظˆط­ط¯ط©') AS selected_unit,
                COALESCE({sales_line_total_expr}, 0) AS net_amount,
                COALESCE({sales_vat_amount_expr}, 0) AS vat_amount,
                COALESCE({sales_line_total_expr},0) + COALESCE({sales_vat_amount_expr},0) AS grand_total
            FROM sales_invoices s
            JOIN sales_invoice_lines l ON l.invoice_id=s.id
            LEFT JOIN products p ON p.id=l.product_id
            LEFT JOIN customers c ON c.id=s.customer_id
            WHERE {sales_where}

            UNION ALL

            SELECT
                p.date,
                'ظ…ط´طھط±ظٹط§طھ' AS doc_type,
                {purchase_invoice_no_expr} AS doc_no,
                COALESCE(sup.name, 'ظ…ظˆط±ط¯ ظ†ظ‚ط¯ظٹ') AS party_name,
                {supplier_tax_expr} AS tax_number,
                COALESCE(pr.name, '') AS item_name,
                {purchase_qty_expr} AS qty,
                {purchase_price_expr} AS unit_price,
                COALESCE({purchase_unit_expr}, 'ظˆط­ط¯ط©') AS selected_unit,
                COALESCE({purchase_line_total_expr},0) AS net_amount,
                COALESCE({purchase_vat_amount_expr}, 0) AS vat_amount,
                COALESCE({purchase_line_total_expr},0) + COALESCE({purchase_vat_amount_expr},0) AS grand_total
            FROM purchase_invoices p
            JOIN purchase_invoice_lines l ON l.invoice_id=p.id
            LEFT JOIN products pr ON pr.id=l.product_id
            LEFT JOIN suppliers sup ON sup.id=p.supplier_id
            WHERE {purchases_where}
            ORDER BY 1 DESC, 3 DESC
            """,
            params_sales + params_purchases,
        )
        rows = cur.fetchall()
        conn.close()

        headers = ["ط§ظ„طھط§ط±ظٹط®", "ط§ظ„ظ†ظˆط¹", "ط±ظ‚ظ… ط§ظ„ظ…ط³طھظ†ط¯", "ط§ظ„ط¬ظ‡ط©", "ط§ظ„ط±ظ‚ظ… ط§ظ„ط¶ط±ظٹط¨ظٹ", "ط§ظ„طµظ†ظپ", "ط§ظ„ظƒظ…ظٹط©", "ط³ط¹ط± ط§ظ„ظˆط­ط¯ط©", "ط§ظ„ظˆط­ط¯ط©", "ط§ظ„طµط§ظپظٹ", "ط§ظ„ط¶ط±ظٹط¨ط©", "ط§ظ„ط¥ط¬ظ…ط§ظ„ظٹ"]
        if request.args.get("format") == "excel":
            return _financial_report_excel("vat_report.xlsx", headers, rows, "طھظ‚ط±ظٹط± ط¶ط±ظٹط¨ط© ط§ظ„ظ‚ظٹظ…ط© ط§ظ„ظ…ط¶ط§ظپط©")

        total_net = sum(float(row[9] or 0) for row in rows)
        total_vat = sum(float(row[10] or 0) for row in rows)
        total_grand = sum(float(row[11] or 0) for row in rows)
        output_vat = sum(float(row[10] or 0) for row in rows if row[1] == "ظ…ط¨ظٹط¹ط§طھ")
        input_vat = sum(float(row[10] or 0) for row in rows if row[1] == "ظ…ط´طھط±ظٹط§طھ")
        net_due = output_vat - input_vat

        return render_template_string(
            """
            {% extends "layout.html" %}
            {% block content %}
            <div class="container-fluid" dir="rtl">
                <div class="d-flex justify-content-between align-items-center mb-3">
                    <div>
                        <h2 class="mb-1">طھظ‚ط±ظٹط± ط¶ط±ظٹط¨ط© ط§ظ„ظ‚ظٹظ…ط© ط§ظ„ظ…ط¶ط§ظپط©</h2>
                        <div class="text-muted">ظٹط¹ط±ط¶ ط§ظ„ظپظˆط§طھظٹط± ط§ظ„ظ…ط±ط­ظ„ط© ظ…ط¹ ط§ظ„ظƒظ…ظٹط© ظˆط³ط¹ط± ط§ظ„ظˆط­ط¯ط© ظˆط§ظ„ظˆط­ط¯ط© ظˆط§ظ„ط±ظ‚ظ… ط§ظ„ط¶ط±ظٹط¨ظٹ.</div>
                    </div>
                    <div class="d-flex gap-2">
                        <a class="btn btn-success" href="{{ url_for('vat_report', date_from=date_from, date_to=date_to, format='excel') }}">طھط­ظ…ظٹظ„ Excel</a>
                        <button class="btn btn-outline-secondary" onclick="window.print()">ط·ط¨ط§ط¹ط©</button>
                    </div>
                </div>

                <form method="get" class="card p-3 mb-3">
                    <div class="row g-2 align-items-end">
                        <div class="col-md-3">
                            <label class="form-label">ظ…ظ† طھط§ط±ظٹط®</label>
                            <input class="form-control" type="date" name="date_from" value="{{ date_from }}">
                        </div>
                        <div class="col-md-3">
                            <label class="form-label">ط¥ظ„ظ‰ طھط§ط±ظٹط®</label>
                            <input class="form-control" type="date" name="date_to" value="{{ date_to }}">
                        </div>
                        <div class="col-md-6 d-flex gap-2">
                            <button class="btn btn-primary">طھط·ط¨ظٹظ‚</button>
                            <a class="btn btn-light border" href="{{ url_for('vat_report') }}">ط¥ظ„ط؛ط§ط،</a>
                        </div>
                    </div>
                </form>

                <div class="row g-3 mb-3">
                    <div class="col-md-4"><div class="card p-3"><div class="text-muted">ط¶ط±ظٹط¨ط© ط§ظ„ظ…ط®ط±ط¬ط§طھ</div><div class="fs-4 fw-bold">{{ '%.2f'|format(output_vat) }}</div></div></div>
                    <div class="col-md-4"><div class="card p-3"><div class="text-muted">ط¶ط±ظٹط¨ط© ط§ظ„ظ…ط¯ط®ظ„ط§طھ</div><div class="fs-4 fw-bold">{{ '%.2f'|format(input_vat) }}</div></div></div>
                    <div class="col-md-4"><div class="card p-3"><div class="text-muted">ط§ظ„طµط§ظپظٹ ط§ظ„ظ…ط³طھط­ظ‚</div><div class="fs-4 fw-bold">{{ '%.2f'|format(net_due) }}</div></div></div>
                </div>

                <div class="table-responsive card p-2">
                    <table class="table table-bordered table-striped align-middle mb-0">
                        <thead><tr>{% for header in headers %}<th>{{ header }}</th>{% endfor %}</tr></thead>
                        <tbody>
                            {% for row in rows %}
                            <tr>
                                <td>{{ row[0] }}</td>
                                <td>{{ row[1] }}</td>
                                <td dir="ltr">{{ row[2] }}</td>
                                <td>{{ row[3] }}</td>
                                <td>{{ row[4] }}</td>
                                <td>{{ row[5] }}</td>
                                <td>{{ '%.2f'|format(row[6] or 0) }}</td>
                                <td>{{ '%.2f'|format(row[7] or 0) }}</td>
                                <td>{{ row[8] }}</td>
                                <td>{{ '%.2f'|format(row[9] or 0) }}</td>
                                <td>{{ '%.2f'|format(row[10] or 0) }}</td>
                                <td>{{ '%.2f'|format(row[11] or 0) }}</td>
                            </tr>
                            {% else %}
                            <tr><td colspan="{{ headers|length }}" class="text-center text-muted">ظ„ط§ طھظˆط¬ط¯ ط¨ظٹط§ظ†ط§طھ ط¶ظ…ظ† ط§ظ„ظپطھط±ط© ط§ظ„ظ…ط­ط¯ط¯ط©.</td></tr>
                            {% endfor %}
                        </tbody>
                        <tfoot>
                            <tr>
                                <th colspan="9">ط§ظ„ط¥ط¬ظ…ط§ظ„ظٹ</th>
                                <th>{{ '%.2f'|format(total_net) }}</th>
                                <th>{{ '%.2f'|format(total_vat) }}</th>
                                <th>{{ '%.2f'|format(total_grand) }}</th>
                            </tr>
                        </tfoot>
                    </table>
                </div>
            </div>
            {% endblock %}
            """,
            rows=rows,
            headers=headers,
            date_from=date_from,
            date_to=date_to,
            total_net=total_net,
            total_vat=total_vat,
            total_grand=total_grand,
            output_vat=output_vat,
            input_vat=input_vat,
            net_due=net_due,
        )

    return vat_report


def build_withholding_tax_report_view(deps):
    db = deps["db"]

    def withholding_tax_report():
        conn = db()
        cur = conn.cursor()
        date_from = (request.args.get("date_from") or "").strip()
        date_to = (request.args.get("date_to") or "").strip()
        customer_tax_expr = _report_tax_number_expr(cur, "customers", "c", "'ط؛ظٹط± ظ…ط³ط¬ظ„'")

        conditions = [
            "si.status='posted'",
            "COALESCE(sil.withholding_applicable, sil.withholding_enabled, 0) = 1",
            "COALESCE(sil.withholding_amount, 0) > 0",
        ]
        params = []
        if date_from:
            conditions.append("si.date >= ?")
            params.append(date_from)
        if date_to:
            conditions.append("si.date <= ?")
            params.append(date_to)
        where_sql = " AND ".join(conditions)

        cur.execute(
            f"""
            SELECT
                COALESCE(NULLIF(si.invoice_number, ''), si.doc_no) AS invoice_number,
                si.date,
                COALESCE(c.name, 'ط¹ظ…ظٹظ„ ظ†ظ‚ط¯ظٹ') AS customer_name,
                {customer_tax_expr} AS customer_tax_number,
                COALESCE(p.name, '') AS items,
                COALESCE(NULLIF(sil.selected_unit, ''), NULLIF(sil.unit_name, ''), p.unit, 'ظˆط­ط¯ط©') AS units,
                COALESCE(sil.subtotal, sil.total, 0) AS invoice_total,
                COALESCE(sil.withholding_rate, 0) AS withholding_rate,
                COALESCE(sil.withholding_amount, 0) AS withholding_amount,
                COALESCE(sil.line_net, sil.net_total, sil.grand_total, sil.total + COALESCE(sil.vat_amount, 0) - COALESCE(sil.withholding_amount, 0)) AS net_amount
            FROM sales_invoices si
            JOIN sales_invoice_lines sil ON sil.invoice_id = si.id
            LEFT JOIN products p ON p.id = sil.product_id
            LEFT JOIN customers c ON c.id = si.customer_id
            WHERE {where_sql}
            ORDER BY si.date DESC, si.id DESC, sil.id DESC
            """,
            params,
        )
        rows = cur.fetchall()
        conn.close()

        headers = ["ط±ظ‚ظ… ط§ظ„ظپط§طھظˆط±ط©", "ط§ظ„طھط§ط±ظٹط®", "ط§ظ„ط¹ظ…ظٹظ„", "ط§ظ„ط±ظ‚ظ… ط§ظ„ط¶ط±ظٹط¨ظٹ", "ط§ظ„ط£طµظ†ط§ظپ", "ط§ظ„ظˆط­ط¯ط©", "ط¥ط¬ظ…ط§ظ„ظٹ ط§ظ„ظپط§طھظˆط±ط©", "ط§ظ„ظ†ط³ط¨ط©", "ظ‚ظٹظ…ط© ط§ظ„ط®طµظ…/ط§ظ„ط¥ط¶ط§ظپط©", "ط§ظ„طµط§ظپظٹ"]
        if request.args.get("format") == "excel":
            return _financial_report_excel("withholding_tax_report.xlsx", headers, rows, "طھظ‚ط±ظٹط± ط¶ط±ظٹط¨ط© ط§ظ„ط®طµظ… ظˆط§ظ„ط¥ط¶ط§ظپط©")

        total_invoice = sum(float(row[6] or 0) for row in rows)
        total_withholding = sum(float(row[8] or 0) for row in rows)
        total_net = sum(float(row[9] or 0) for row in rows)
        return render_template_string(
            """
            {% extends "layout.html" %}
            {% block content %}
            <div class="container-fluid" dir="rtl">
                <div class="d-flex justify-content-between align-items-center mb-3">
                    <div>
                        <h2 class="mb-1">طھظ‚ط±ظٹط± ط§ظ„ط®طµظ… ظˆط§ظ„ط¥ط¶ط§ظپط©</h2>
                        <div class="text-muted">ظپظˆط§طھظٹط± ط§ظ„ط¨ظٹط¹ ط§ظ„ظ…ط±ط­ظ„ط© ط§ظ„ط®ط§ط¶ط¹ط© ظ„ظ„ط®طµظ… ظˆط§ظ„ط¥ط¶ط§ظپط©.</div>
                    </div>
                    <a class="btn btn-success" href="{{ url_for('withholding_tax_report', date_from=date_from, date_to=date_to, format='excel') }}">طھط­ظ…ظٹظ„ Excel</a>
                </div>

                <form method="get" class="card p-3 mb-3">
                    <div class="row g-2 align-items-end">
                        <div class="col-md-3"><label class="form-label">ظ…ظ† طھط§ط±ظٹط®</label><input class="form-control" type="date" name="date_from" value="{{ date_from }}"></div>
                        <div class="col-md-3"><label class="form-label">ط¥ظ„ظ‰ طھط§ط±ظٹط®</label><input class="form-control" type="date" name="date_to" value="{{ date_to }}"></div>
                        <div class="col-md-6 d-flex gap-2">
                            <button class="btn btn-primary">طھط·ط¨ظٹظ‚</button>
                            <a class="btn btn-light border" href="{{ url_for('withholding_tax_report') }}">ط¥ظ„ط؛ط§ط،</a>
                        </div>
                    </div>
                </form>

                <div class="row g-3 mb-3">
                    <div class="col-md-4"><div class="card p-3"><div class="text-muted">ط¥ط¬ظ…ط§ظ„ظٹ ط§ظ„ظپظˆط§طھظٹط±</div><div class="fs-4 fw-bold">{{ '%.2f'|format(total_invoice) }}</div></div></div>
                    <div class="col-md-4"><div class="card p-3"><div class="text-muted">ط¥ط¬ظ…ط§ظ„ظٹ ط§ظ„ط®طµظ…/ط§ظ„ط¥ط¶ط§ظپط©</div><div class="fs-4 fw-bold">{{ '%.2f'|format(total_withholding) }}</div></div></div>
                    <div class="col-md-4"><div class="card p-3"><div class="text-muted">ط§ظ„طµط§ظپظٹ</div><div class="fs-4 fw-bold">{{ '%.2f'|format(total_net) }}</div></div></div>
                </div>

                <div class="table-responsive card p-2">
                    <table class="table table-bordered table-striped align-middle mb-0">
                        <thead><tr>{% for header in headers %}<th>{{ header }}</th>{% endfor %}</tr></thead>
                        <tbody>
                            {% for row in rows %}
                            <tr>
                                <td dir="ltr">{{ row[0] }}</td><td>{{ row[1] }}</td><td>{{ row[2] }}</td><td>{{ row[3] }}</td><td>{{ row[4] }}</td><td>{{ row[5] }}</td>
                                <td>{{ '%.2f'|format(row[6] or 0) }}</td><td>{{ '%.2f'|format(row[7] or 0) }}</td><td>{{ '%.2f'|format(row[8] or 0) }}</td><td>{{ '%.2f'|format(row[9] or 0) }}</td>
                            </tr>
                            {% else %}
                            <tr><td colspan="{{ headers|length }}" class="text-center text-muted">ظ„ط§ طھظˆط¬ط¯ ظپظˆط§طھظٹط± ط¶ظ…ظ† ط§ظ„ظپطھط±ط© ط§ظ„ظ…ط­ط¯ط¯ط©.</td></tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            </div>
            {% endblock %}
            """,
            rows=rows,
            headers=headers,
            date_from=date_from,
            date_to=date_to,
            total_invoice=total_invoice,
            total_withholding=total_withholding,
            total_net=total_net,
        )

    return withholding_tax_report

