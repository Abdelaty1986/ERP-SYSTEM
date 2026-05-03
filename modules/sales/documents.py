from io import BytesIO

from flask import flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def _customer_tax_number_expr(customer_columns):
    for column_name in ("tax_registration_number", "tax_number", "tax_id"):
        if column_name in customer_columns:
            return f"COALESCE(NULLIF(c.{column_name}, ''), 'غير مسجل')"
    return "'غير مسجل'"


def _invoice_company_party(cur, invoice_id, invoice_kind):
    if invoice_kind == "sale":
        cur.execute(
            """
            SELECT s.id,s.doc_no,s.date,COALESCE(c.name,'عميل نقدي'),COALESCE(c.tax_id,''),s.grand_total,s.tax_amount,s.withholding_amount
            FROM sales_invoices s
            LEFT JOIN customers c ON c.id=s.customer_id
            WHERE s.id=?
            """,
            (invoice_id,),
        )
    else:
        cur.execute(
            """
            SELECT p.id,p.doc_no,p.date,COALESCE(s.name,'مورد نقدي'),COALESCE(s.tax_id,''),p.grand_total,p.tax_amount,p.withholding_amount
            FROM purchase_invoices p
            LEFT JOIN suppliers s ON s.id=p.supplier_id
            WHERE p.id=?
            """,
            (invoice_id,),
        )
    return cur.fetchone()


def _invoice_lines(cur, invoice_id, invoice_kind):
    if invoice_kind == "sale":
        cur.execute(
            """
            SELECT p.code,p.name,COALESCE(NULLIF(l.selected_unit,''), NULLIF(l.unit_name,''), p.unit, 'وحدة'),
                   COALESCE(l.qty,l.quantity),l.unit_price,l.total,
                   COALESCE(l.vat_enabled,1),COALESCE(l.vat_amount,0),
                   COALESCE(l.withholding_enabled,0),COALESCE(l.withholding_amount,0),
                   COALESCE(l.grand_total,l.total + COALESCE(l.vat_amount,0))
            FROM sales_invoice_lines l
            JOIN products p ON p.id=l.product_id
            WHERE l.invoice_id=?
            ORDER BY l.id
            """,
            (invoice_id,),
        )
        rows = cur.fetchall()
        if rows:
            return rows
        cur.execute(
            """
            SELECT p.code,p.name,p.unit,s.quantity,s.unit_price,s.total,
                   1,s.tax_amount,CASE WHEN COALESCE(s.withholding_rate,0) > 0 THEN 1 ELSE 0 END,
                   COALESCE(s.withholding_amount,0),s.grand_total
            FROM sales_invoices s
            JOIN products p ON p.id=s.product_id
            WHERE s.id=?
            """,
            (invoice_id,),
        )
        return cur.fetchall()
    cur.execute(
        """
        SELECT p.code,p.name,COALESCE(NULLIF(l.selected_unit,''), NULLIF(l.unit_name,''), p.unit, 'وحدة'),
               COALESCE(l.qty,l.quantity),l.unit_price,l.total,
               COALESCE(l.vat_enabled,1),COALESCE(l.vat_amount,0),
               COALESCE(l.withholding_enabled,0),COALESCE(l.withholding_amount,0),
               COALESCE(l.grand_total,l.total + COALESCE(l.vat_amount,0))
        FROM purchase_invoice_lines l
        JOIN products p ON p.id=l.product_id
        WHERE l.invoice_id=?
        ORDER BY l.id
        """,
        (invoice_id,),
    )
    rows = cur.fetchall()
    if rows:
        return rows
    cur.execute(
        """
        SELECT p.code,p.name,p.unit,s.quantity,s.unit_price,s.total,
               1,s.tax_amount,CASE WHEN COALESCE(s.withholding_rate,0) > 0 THEN 1 ELSE 0 END,
               COALESCE(s.withholding_amount,0),s.grand_total
        FROM purchase_invoices s
        JOIN products p ON p.id=s.product_id
        WHERE s.id=?
        """,
        (invoice_id,),
    )
    return cur.fetchall()


def _build_invoice_workbook(company, header, lines, party_label):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.sheet_view.rightToLeft = True
    ws["A1"] = company.get("name") or "LedgerX-SYSTEM"
    ws["A2"] = f"{party_label}: {header[3]}"
    ws["A3"] = f"رقم الفاتورة: {header[1]}"
    ws["A4"] = f"تاريخ الفاتورة: {header[2]}"
    ws["A5"] = f"الرقم الضريبي: {header[4] or '-'}"
    for cell in ("A1", "A2", "A3", "A4", "A5"):
        ws[cell].font = Font(bold=True)
    headers = [
        "كود الصنف",
        "اسم الصنف",
        "الوحدة",
        "الكمية",
        "سعر الوحدة",
        "الإجمالي قبل الضريبة",
        "خاضع VAT",
        "قيمة VAT",
        "خاضع خصم وإضافة",
        "قيمة خصم وإضافة",
        "الإجمالي النهائي",
    ]
    row_idx = 7
    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=idx, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    total = vat_total = withholding_total = 0
    for line in lines:
        row_idx += 1
        total += line[5] or 0
        vat_total += line[7] or 0
        withholding_total += line[9] or 0
        values = [
            line[0] or "",
            line[1],
            line[2] or "",
            line[3],
            line[4],
            line[5],
            "نعم" if line[6] else "لا",
            line[7],
            "نعم" if line[8] else "لا",
            line[9],
            line[10],
        ]
        for idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=idx, value=value)
    row_idx += 2
    ws.cell(row=row_idx, column=1, value="إجمالي قبل الضريبة").font = Font(bold=True)
    ws.cell(row=row_idx, column=2, value=total)
    ws.cell(row=row_idx + 1, column=1, value="إجمالي VAT").font = Font(bold=True)
    ws.cell(row=row_idx + 1, column=2, value=vat_total)
    ws.cell(row=row_idx + 2, column=1, value="إجمالي خصم وإضافة").font = Font(bold=True)
    ws.cell(row=row_idx + 2, column=2, value=withholding_total)
    ws.cell(row=row_idx + 3, column=1, value="إجمالي الفاتورة").font = Font(bold=True)
    ws.cell(row=row_idx + 3, column=2, value=header[5] or 0)
    widths = [16, 28, 14, 12, 14, 18, 12, 14, 18, 16, 18]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    return wb


def build_print_sale_view(deps):
    db = deps["db"]
    get_company_settings = deps["get_company_settings"]
    amount_to_words = deps["amount_to_words"]

    def print_sale(id):
        conn = db()
        cur = conn.cursor()
        company = get_company_settings(cur)
        cur.execute(
            """
            SELECT s.id,s.date,COALESCE(c.name,'بيع نقدي'),COALESCE(c.phone,''),COALESCE(c.address,''),
                   s.quantity,s.unit_price,s.total,s.tax_rate,s.tax_amount,s.withholding_rate,s.withholding_amount,s.grand_total,
                   s.payment_type,s.status,s.cancel_reason,s.due_date
            FROM sales_invoices s
            LEFT JOIN customers c ON s.customer_id=c.id
            WHERE s.id=?
            """,
            (id,),
        )
        doc = cur.fetchone()
        lines = _invoice_lines(cur, id, "sale") if doc else []
        conn.close()
        if not doc:
            flash("فاتورة البيع غير موجودة.", "danger")
            return redirect(url_for("sales"))
        return render_template(
            "print_document.html",
            company=company,
            doc=doc,
            lines=lines,
            doc_type="فاتورة بيع",
            party_label="العميل",
            sales_invoice=True,
            amount_in_words=amount_to_words(doc[12]),
        )

    return print_sale


def build_print_purchase_view(deps):
    db = deps["db"]
    get_company_settings = deps["get_company_settings"]
    amount_to_words = deps["amount_to_words"]

    def print_purchase(id):
        conn = db()
        cur = conn.cursor()
        company = get_company_settings(cur)
        cur.execute(
            """
            SELECT p.id,p.date,COALESCE(s.name,'شراء نقدي'),COALESCE(s.phone,''),COALESCE(s.address,''),
                   p.quantity,p.unit_price,p.total,p.tax_rate,p.tax_amount,p.withholding_rate,p.withholding_amount,p.grand_total,
                   p.payment_type,p.status,p.cancel_reason,p.supplier_invoice_no,p.supplier_invoice_date,p.due_date
            FROM purchase_invoices p
            LEFT JOIN suppliers s ON p.supplier_id=s.id
            WHERE p.id=?
            """,
            (id,),
        )
        doc = cur.fetchone()
        lines = _invoice_lines(cur, id, "purchase") if doc else []
        conn.close()
        if not doc:
            flash("فاتورة المورد غير موجودة.", "danger")
            return redirect(url_for("purchases"))
        return render_template(
            "print_document.html",
            company=company,
            doc=doc,
            lines=lines,
            doc_type="تسجيل فاتورة مورد",
            party_label="المورد",
            supplier_invoice=True,
            amount_in_words=amount_to_words(doc[12]),
        )

    return print_purchase


def build_print_sales_credit_note_view(deps):
    db = deps["db"]
    get_company_settings = deps["get_company_settings"]
    amount_to_words = deps["amount_to_words"]

    def print_sales_credit_note(id):
        conn = db()
        cur = conn.cursor()
        company = get_company_settings(cur)
        cur.execute(
            """
            SELECT scn.id,scn.date,scn.doc_no,COALESCE(c.name,'عميل نقدي'),COALESCE(c.phone,''),COALESCE(c.address,''),
                   p.name,p.unit,scn.quantity,scn.unit_price,scn.total,scn.tax_amount,scn.grand_total,
                   scn.notes,scn.sales_return_id,scn.sales_invoice_id
            FROM sales_credit_notes scn
            LEFT JOIN customers c ON c.id=scn.customer_id
            LEFT JOIN products p ON p.id=scn.product_id
            WHERE scn.id=?
            """,
            (id,),
        )
        doc = cur.fetchone()
        conn.close()
        if not doc:
            flash("إشعار التسوية الدائن غير موجود.", "danger")
            return redirect(url_for("sales_credit_notes"))
        return render_template(
            "print_customer_note.html",
            company=company,
            doc=doc,
            doc_title="إشعار تسوية دائن للعميل",
            note_kind="credit",
            party_label="العميل",
            amount_in_words=amount_to_words(doc[12]),
            source_label=f"من واقع مردودات المبيعات رقم {doc[14]} / الفاتورة الأصلية رقم {doc[15]}",
        )

    return print_sales_credit_note


def build_print_supplier_debit_note_view(deps):
    db = deps["db"]
    get_company_settings = deps["get_company_settings"]
    amount_to_words = deps["amount_to_words"]

    def print_supplier_debit_note(id):
        conn = db()
        cur = conn.cursor()
        company = get_company_settings(cur)
        cur.execute(
            """
            SELECT sdn.id,sdn.date,sdn.doc_no,COALESCE(s.name,'مورد نقدي'),COALESCE(s.phone,''),COALESCE(s.address,''),
                   p.name,p.unit,sdn.quantity,sdn.unit_price,sdn.total,sdn.tax_amount,sdn.grand_total,
                   sdn.notes,sdn.purchase_return_id,sdn.purchase_invoice_id
            FROM supplier_debit_notes sdn
            LEFT JOIN suppliers s ON s.id=sdn.supplier_id
            LEFT JOIN products p ON p.id=sdn.product_id
            WHERE sdn.id=?
            """,
            (id,),
        )
        doc = cur.fetchone()
        conn.close()
        if not doc:
            flash("إشعار التسوية المدين غير موجود.", "danger")
            return redirect(url_for("supplier_debit_notes"))
        return render_template(
            "print_customer_note.html",
            company=company,
            doc=doc,
            doc_title="إشعار تسوية مدين للمورد",
            note_kind="debit",
            party_label="المورد",
            amount_in_words=amount_to_words(doc[12]),
            source_label=f"من واقع مردودات المشتريات رقم {doc[14]} / الفاتورة الأصلية رقم {doc[15]}",
        )

    return print_supplier_debit_note


def build_prepare_sales_credit_note_einvoice_view(deps):
    db = deps["db"]
    prepare_einvoice_document = deps["prepare_einvoice_document"]
    log_action = deps["log_action"]

    def prepare_sales_credit_note_einvoice(id):
        conn = db()
        cur = conn.cursor()
        cur.execute("SELECT id FROM sales_credit_notes WHERE id=?", (id,))
        if not cur.fetchone():
            conn.close()
            flash("إشعار التسوية الدائن غير موجود.", "danger")
            return redirect(url_for("sales_credit_notes"))
        _, created = prepare_einvoice_document(cur, "sales_credit_note", id)
        log_action(cur, "prepare", "e_invoice_documents", None, f"sales_credit_note={id}")
        conn.commit()
        conn.close()
        flash("تم تجهيز إشعار التسوية الدائن للرفع على بوابة الضرائب." if created else "إشعار التسوية الدائن مجهز بالفعل للرفع.", "success")
        return redirect(url_for("sales_credit_notes"))

    return prepare_sales_credit_note_einvoice


def build_export_sale_excel_view(deps):
    db = deps["db"]
    get_company_settings = deps["get_company_settings"]

    def export_sale_excel(id):
        conn = db()
        cur = conn.cursor()
        company = get_company_settings(cur)
        header = _invoice_company_party(cur, id, "sale")
        if not header:
            conn.close()
            flash("فاتورة البيع غير موجودة.", "danger")
            return redirect(url_for("sales"))
        lines = _invoice_lines(cur, id, "sale")
        conn.close()
        wb = _build_invoice_workbook(company, header, lines, "العميل")
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name=f"{header[1] or 'sale-invoice'}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return export_sale_excel


def build_export_purchase_excel_view(deps):
    db = deps["db"]
    get_company_settings = deps["get_company_settings"]

    def export_purchase_excel(id):
        conn = db()
        cur = conn.cursor()
        company = get_company_settings(cur)
        header = _invoice_company_party(cur, id, "purchase")
        if not header:
            conn.close()
            flash("فاتورة المورد غير موجودة.", "danger")
            return redirect(url_for("purchases"))
        lines = _invoice_lines(cur, id, "purchase")
        conn.close()
        wb = _build_invoice_workbook(company, header, lines, "المورد")
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name=f"{header[1] or 'purchase-invoice'}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return export_purchase_excel


def build_export_sales_invoices_excel_view(deps):
    db = deps["db"]

    def export_sales_invoices_excel():
        from_date = (request.args.get("from_date") or "").strip()
        to_date = (request.args.get("to_date") or "").strip()
        if not from_date or not to_date:
            flash("اختر من تاريخ وإلى تاريخ قبل تحميل ملف Excel.", "warning")
            return redirect(url_for("sales_invoices"))

        conn = db()
        cur = conn.cursor()
        sales_invoice_columns = {row[1] for row in cur.execute("PRAGMA table_info(sales_invoices)").fetchall()}
        user_columns = {row[1] for row in cur.execute("PRAGMA table_info(users)").fetchall()} if cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name='users'").fetchone() else set()

        created_by_expr = "''"
        join_users_sql = ""
        if "created_by" in sales_invoice_columns:
            if "username" in user_columns:
                join_users_sql = "LEFT JOIN users u ON CAST(u.id AS TEXT)=CAST(si.created_by AS TEXT) OR u.username=si.created_by"
                created_by_expr = "COALESCE(u.username, si.created_by, '')"
            else:
                created_by_expr = "COALESCE(si.created_by, '')"

        cur.execute(
            f"""
            SELECT
                si.doc_no,
                si.date,
                COALESCE(c.name, 'عميل نقدي') AS customer_name,
                COALESCE(si.payment_type, '') AS payment_type,
                COALESCE(si.total, 0) AS subtotal,
                COALESCE(si.withholding_amount, 0) AS discount_amount,
                COALESCE(si.tax_amount, 0) AS tax_amount,
                COALESCE(si.grand_total, 0) AS net_amount,
                COALESCE(si.status, 'draft') AS status,
                {created_by_expr} AS created_by_name
            FROM sales_invoices si
            LEFT JOIN customers c ON si.customer_id = c.id
            {join_users_sql}
            WHERE si.date >= ? AND si.date <= ?
            ORDER BY si.id DESC
            """,
            (from_date, to_date),
        )
        rows = cur.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Invoices"
        ws.sheet_view.rightToLeft = True
        headers = [
            "رقم الفاتورة",
            "تاريخ الفاتورة",
            "اسم العميل",
            "نوع الدفع",
            "الإجمالي قبل الخصم",
            "الخصم",
            "الضريبة",
            "الصافي",
            "حالة الترحيل",
            "اسم المستخدم",
        ]
        ws.append(headers)
        for idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        payment_labels = {"cash": "نقدي", "credit": "آجل"}
        status_labels = {"draft": "غير مرحلة", "posted": "مرحلة", "cancelled": "ملغاة"}
        for row in rows:
            ws.append(
                [
                    row[0] or "",
                    row[1] or "",
                    row[2] or "عميل نقدي",
                    payment_labels.get(row[3], row[3] or ""),
                    float(row[4] or 0),
                    float(row[5] or 0),
                    float(row[6] or 0),
                    float(row[7] or 0),
                    status_labels.get(row[8], row[8] or ""),
                    row[9] or "",
                ]
            )
        for column, width in zip("ABCDEFGHIJ", [18, 16, 28, 14, 18, 14, 14, 16, 16, 20]):
            ws.column_dimensions[column].width = width
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(
            out,
            as_attachment=True,
            download_name=f"sales_invoices_{from_date}_to_{to_date}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return export_sales_invoices_excel
