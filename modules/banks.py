import sqlite3
from datetime import date, datetime
from io import BytesIO

from flask import flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font

from modules.accounting.ledger_engine import post_simple_entry


BANK_SYSTEM_ACCOUNTS = {
    "receivable_checks": ("1195", "أوراق قبض", "أصول"),
    "payable_checks": ("2195", "أوراق دفع", "خصوم"),
    "bank_commission": ("5295", "عمولات بنكية", "مصروفات"),
    "bank_interest": ("4295", "فوائد بنكية", "إيرادات"),
    "reconciliation_diff": ("5395", "فروق تسوية بنكية", "مصروفات"),
    "bank_opening": ("3195", "رصيد افتتاحي للبنوك", "حقوق ملكية"),
}

TRANSACTION_TYPES = [
    ("deposit", "إيداع بنكي"),
    ("withdrawal", "سحب بنكي"),
    ("bank_to_bank_out", "تحويل من بنك إلى بنك"),
    ("treasury_to_bank", "تحويل من خزينة إلى بنك"),
    ("bank_to_treasury", "تحويل من بنك إلى خزينة"),
    ("bank_expense", "مصروف بنكي"),
    ("bank_commission", "عمولة بنك"),
    ("bank_interest", "فوائد بنكية"),
    ("bank_reconciliation", "تسوية بنكية"),
    ("opening_balance", "رصيد افتتاحي للبنك"),
]
RECEIVABLE_STATUSES = {
    "portfolio": "في الحافظة",
    "deposited": "مودع",
    "collected": "محصل",
    "bounced": "مرفوض",
}
PAYABLE_STATUSES = {
    "issued": "مُصدر",
    "delivered": "مُسلم",
    "cashed": "مُصرف",
    "cancelled": "ملغي",
}


def _add_column_if_missing(cur, table_name, column_name, definition):
    cur.execute(f"PRAGMA table_info({table_name})")
    columns = [row[1] for row in cur.fetchall()]
    if column_name not in columns:
        cur.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition}")


def _table_exists(cur, table_name):
    cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
    return cur.fetchone() is not None


def _ensure_bank_tables(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS banks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            branch TEXT,
            account_number TEXT,
            iban TEXT,
            currency TEXT NOT NULL DEFAULT 'EGP',
            opening_balance REAL NOT NULL DEFAULT 0,
            current_balance REAL NOT NULL DEFAULT 0,
            gl_account_id INTEGER,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_by TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS bank_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bank_id INTEGER NOT NULL,
            related_bank_id INTEGER,
            txn_group TEXT,
            txn_type TEXT NOT NULL,
            doc_no TEXT,
            txn_date TEXT NOT NULL,
            description TEXT,
            reference_no TEXT,
            amount REAL NOT NULL DEFAULT 0,
            signed_amount REAL NOT NULL DEFAULT 0,
            counterparty_account_id INTEGER,
            journal_id INTEGER,
            balance_before REAL NOT NULL DEFAULT 0,
            balance_after REAL NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'posted',
            notes TEXT,
            source_type TEXT,
            source_id INTEGER,
            created_by TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS bank_reconciliations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bank_id INTEGER NOT NULL,
            doc_no TEXT,
            statement_date TEXT NOT NULL,
            from_date TEXT,
            to_date TEXT,
            opening_balance REAL NOT NULL DEFAULT 0,
            system_balance REAL NOT NULL DEFAULT 0,
            statement_balance REAL NOT NULL DEFAULT 0,
            difference_amount REAL NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'draft',
            notes TEXT,
            journal_id INTEGER,
            created_by TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS bank_reconciliation_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reconciliation_id INTEGER NOT NULL,
            line_date TEXT,
            description TEXT,
            reference_no TEXT,
            amount REAL NOT NULL DEFAULT 0,
            matched_transaction_id INTEGER,
            match_status TEXT NOT NULL DEFAULT 'unmatched',
            difference_amount REAL NOT NULL DEFAULT 0,
            notes TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS receivable_checks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_no TEXT,
            check_number TEXT NOT NULL,
            bank_id INTEGER NOT NULL,
            customer_id INTEGER NOT NULL,
            due_date TEXT NOT NULL,
            amount REAL NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'portfolio',
            notes TEXT,
            received_journal_id INTEGER,
            deposit_journal_id INTEGER,
            bounce_journal_id INTEGER,
            bank_transaction_id INTEGER,
            created_by TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS payable_checks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_no TEXT,
            check_number TEXT NOT NULL,
            bank_id INTEGER NOT NULL,
            supplier_id INTEGER NOT NULL,
            due_date TEXT NOT NULL,
            amount REAL NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'issued',
            notes TEXT,
            issue_journal_id INTEGER,
            cash_journal_id INTEGER,
            cancel_journal_id INTEGER,
            bank_transaction_id INTEGER,
            created_by TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS ux_bank_transactions_type_doc
        ON bank_transactions(txn_type, doc_no)
        WHERE NULLIF(TRIM(COALESCE(doc_no, '')), '') IS NOT NULL
        """
    )
    cur.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS ux_receivable_checks_bank_number
        ON receivable_checks(bank_id, check_number)
        """
    )
    cur.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS ux_payable_checks_bank_number
        ON payable_checks(bank_id, check_number)
        """
    )
    cur.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS ux_bank_reconciliations_doc_no
        ON bank_reconciliations(doc_no)
        WHERE NULLIF(TRIM(COALESCE(doc_no, '')), '') IS NOT NULL
        """
    )
    for table_name in ("banks", "bank_transactions", "bank_reconciliations", "receivable_checks", "payable_checks"):
        _add_column_if_missing(cur, table_name, "created_by", "TEXT")


def _ensure_sequence_rows(cur):
    defaults = [
        ("banks", "BNK", 1),
        ("bank_transactions", "BTX", 1),
        ("bank_transfers", "BTR", 1),
        ("bank_reconciliations", "BRC", 1),
        ("receivable_checks", "RCH", 1),
        ("payable_checks", "PCH", 1),
    ]
    for doc_type, prefix, next_number in defaults:
        cur.execute(
            "INSERT OR IGNORE INTO document_sequences(doc_type,prefix,next_number) VALUES (?,?,?)",
            (doc_type, prefix, next_number),
        )


def _ensure_system_accounts(cur):
    for code, name, account_type in BANK_SYSTEM_ACCOUNTS.values():
        cur.execute("SELECT id FROM accounts WHERE code=?", (code,))
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO accounts(code,name,type) VALUES (?,?,?)",
                (code, name, account_type),
            )


def _safe_float(value):
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _post_entry_by_ids(cur, txn_date, description, debit_account_id, credit_account_id, amount, source_type):
    amount = _safe_float(amount)
    if amount <= 0:
        return None
    cur.execute("SELECT code FROM accounts WHERE id=?", (debit_account_id,))
    debit_row = cur.fetchone()
    cur.execute("SELECT code FROM accounts WHERE id=?", (credit_account_id,))
    credit_row = cur.fetchone()
    if not debit_row or not credit_row:
        raise ValueError("الحساب المحاسبي غير موجود.")
    return post_simple_entry(
        cur,
        txn_date,
        description,
        debit_row[0],
        credit_row[0],
        amount,
        source_type=source_type,
    )


def _system_account_id(cur, key):
    code = BANK_SYSTEM_ACCOUNTS[key][0]
    cur.execute("SELECT id FROM accounts WHERE code=?", (code,))
    row = cur.fetchone()
    return row[0] if row else None


def _account_id_by_code(cur, code):
    cur.execute("SELECT id FROM accounts WHERE code=?", (code,))
    row = cur.fetchone()
    return row[0] if row else None


def _default_treasury_account_id(cur):
    for code in ("1100", "1110", "1120", "1200", "1210"):
        cur.execute("SELECT id FROM accounts WHERE code=?", (code,))
        row = cur.fetchone()
        if row:
            return row[0]
    cur.execute("SELECT id FROM accounts ORDER BY code LIMIT 1")
    row = cur.fetchone()
    return row[0] if row else None


def _fetch_accounts(cur):
    cur.execute("SELECT id, code, name FROM accounts ORDER BY code, name")
    return cur.fetchall()


def _fetch_banks(cur, active_only=False):
    sql = """
        SELECT b.id,b.name,b.branch,b.account_number,b.iban,b.currency,
               b.opening_balance,b.current_balance,b.gl_account_id,b.is_active,
               a.code,a.name
        FROM banks b
        LEFT JOIN accounts a ON a.id=b.gl_account_id
    """
    if active_only:
        sql += " WHERE b.is_active=1"
    sql += " ORDER BY b.name,b.id"
    cur.execute(sql)
    return cur.fetchall()


def _fetch_customers(cur):
    cur.execute("SELECT id, name FROM customers ORDER BY name")
    return cur.fetchall()


def _fetch_suppliers(cur):
    cur.execute("SELECT id, name FROM suppliers ORDER BY name")
    return cur.fetchall()


def _bank_record(cur, bank_id):
    cur.execute(
        """
        SELECT b.id,b.name,b.branch,b.account_number,b.iban,b.currency,
               b.opening_balance,b.current_balance,b.gl_account_id,b.is_active,
               a.code,a.name
        FROM banks b
        LEFT JOIN accounts a ON a.id=b.gl_account_id
        WHERE b.id=?
        """,
        (bank_id,),
    )
    return cur.fetchone()


def _recalculate_bank_balance(cur, bank_id):
    cur.execute("SELECT COALESCE(opening_balance,0) FROM banks WHERE id=?", (bank_id,))
    row = cur.fetchone()
    if not row:
        return 0
    opening_balance = float(row[0] or 0)
    cur.execute(
        """
        SELECT id, COALESCE(signed_amount,0)
        FROM bank_transactions
        WHERE bank_id=? AND status='posted'
        ORDER BY txn_date, id
        """,
        (bank_id,),
    )
    running = opening_balance
    for txn_id, signed_amount in cur.fetchall():
        before = running
        running = round(running + float(signed_amount or 0), 2)
        cur.execute(
            "UPDATE bank_transactions SET balance_before=?, balance_after=? WHERE id=?",
            (before, running, txn_id),
        )
    cur.execute("UPDATE banks SET current_balance=?, updated_at=CURRENT_TIMESTAMP WHERE id=?", (running, bank_id))
    return running


def _statement_opening_balance(cur, bank_id, from_date=None):
    cur.execute("SELECT COALESCE(opening_balance,0) FROM banks WHERE id=?", (bank_id,))
    opening = float((cur.fetchone() or [0])[0] or 0)
    if not from_date:
        return opening
    cur.execute(
        """
        SELECT COALESCE(SUM(signed_amount),0)
        FROM bank_transactions
        WHERE bank_id=? AND status='posted' AND txn_date < ?
        """,
        (bank_id, from_date),
    )
    return round(opening + float((cur.fetchone() or [0])[0] or 0), 2)


def _insert_bank_transaction(
    cur,
    bank_id,
    txn_type,
    doc_no,
    txn_date,
    description,
    amount,
    signed_amount,
    counterparty_account_id=None,
    journal_id=None,
    related_bank_id=None,
    reference_no=None,
    notes="",
    txn_group=None,
    source_type=None,
    source_id=None,
):
    current_balance = _recalculate_bank_balance(cur, bank_id)
    balance_before = current_balance
    balance_after = round(balance_before + float(signed_amount or 0), 2)
    cur.execute(
        """
        INSERT INTO bank_transactions(
            bank_id,related_bank_id,txn_group,txn_type,doc_no,txn_date,description,
            reference_no,amount,signed_amount,counterparty_account_id,journal_id,
            balance_before,balance_after,status,notes,source_type,source_id,created_by
        )
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            bank_id,
            related_bank_id,
            txn_group,
            txn_type,
            doc_no,
            txn_date,
            description,
            reference_no,
            _safe_float(amount),
            round(float(signed_amount or 0), 2),
            counterparty_account_id,
            journal_id,
            balance_before,
            balance_after,
            "posted",
            notes,
            source_type,
            source_id,
            session.get("username", ""),
        ),
    )
    txn_id = cur.lastrowid
    cur.execute("UPDATE banks SET current_balance=?, updated_at=CURRENT_TIMESTAMP WHERE id=?", (balance_after, bank_id))
    return txn_id


def _bank_statement_rows(cur, bank_id, from_date="", to_date="", txn_type=""):
    sql = """
        SELECT bt.id,bt.txn_date,bt.doc_no,bt.txn_type,bt.description,bt.reference_no,
               bt.amount,bt.signed_amount,bt.balance_before,bt.balance_after,
               rb.name,acc.code,acc.name
        FROM bank_transactions bt
        LEFT JOIN banks rb ON rb.id=bt.related_bank_id
        LEFT JOIN accounts acc ON acc.id=bt.counterparty_account_id
        WHERE bt.bank_id=?
    """
    params = [bank_id]
    if from_date:
        sql += " AND bt.txn_date >= ?"
        params.append(from_date)
    if to_date:
        sql += " AND bt.txn_date <= ?"
        params.append(to_date)
    if txn_type:
        sql += " AND bt.txn_type = ?"
        params.append(txn_type)
    sql += " ORDER BY bt.txn_date, bt.id"
    cur.execute(sql, params)
    return cur.fetchall()


def _statement_excel(filename, title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    ws.title = "Banks"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col_index, value=header).font = Font(bold=True)
    row_index = 4
    for row in rows:
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=value)
        row_index += 1
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _parse_reconciliation_lines(text):
    rows = []
    for raw_line in (text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        parts = [part.strip() for part in line.split("|")]
        if len(parts) < 2:
            continue
        line_date = parts[0]
        amount = _safe_float(parts[1])
        reference_no = parts[2] if len(parts) > 2 else ""
        description = parts[3] if len(parts) > 3 else ""
        rows.append((line_date, amount, reference_no, description))
    return rows


def _match_reconciliation_line(cur, bank_id, line_date, amount, reference_no):
    conditions = ["bt.bank_id=?"]
    params = [bank_id]
    if line_date:
        conditions.append("bt.txn_date=?")
        params.append(line_date)
    if amount:
        conditions.append("ABS(bt.signed_amount)=?")
        params.append(abs(_safe_float(amount)))
    if reference_no:
        conditions.append("(bt.reference_no=? OR bt.doc_no=? )")
        params.extend([reference_no, reference_no])
    sql = f"""
        SELECT bt.id,bt.signed_amount,bt.reference_no,bt.doc_no
        FROM bank_transactions bt
        WHERE {' AND '.join(conditions)}
          AND NOT EXISTS (
              SELECT 1
              FROM bank_reconciliation_lines brl
              WHERE brl.matched_transaction_id=bt.id
          )
        ORDER BY bt.txn_date, bt.id
        LIMIT 1
    """
    cur.execute(sql, params)
    return cur.fetchone()


def _build_check_summary(cur, table_name, customer_side=True):
    party_table = "customers" if customer_side else "suppliers"
    party_fk = "customer_id" if customer_side else "supplier_id"
    status_map = RECEIVABLE_STATUSES if customer_side else PAYABLE_STATUSES
    cur.execute(
        f"""
        SELECT c.id,c.doc_no,c.check_number,b.name,p.name,c.due_date,c.amount,c.status
        FROM {table_name} c
        JOIN banks b ON b.id=c.bank_id
        JOIN {party_table} p ON p.id=c.{party_fk}
        ORDER BY c.created_at DESC, c.id DESC
        """
    )
    rows = cur.fetchall()
    due_today = date.today().isoformat()
    due = [row for row in rows if row[5] == due_today]
    overdue = [row for row in rows if row[5] < due_today and row[7] not in ("collected", "cashed", "cancelled")]
    bounced = [row for row in rows if row[7] in ("bounced",)]
    if not customer_side:
        bounced = [row for row in rows if row[7] == "cancelled"]
    return rows, due, overdue, bounced, status_map


def _ensure_context(cur):
    _ensure_bank_tables(cur)
    _ensure_sequence_rows(cur)
    _ensure_system_accounts(cur)


def build_banks_view(deps):
    db = deps["db"]
    parse_positive_amount = deps["parse_positive_amount"]
    log_action = deps["log_action"]

    def banks():
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)

        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            branch = (request.form.get("branch") or "").strip()
            account_number = (request.form.get("account_number") or "").strip()
            iban = (request.form.get("iban") or "").strip()
            currency = (request.form.get("currency") or "EGP").strip() or "EGP"
            opening_balance = parse_positive_amount(request.form.get("opening_balance"))
            gl_account_id = request.form.get("gl_account_id") or None
            is_active = 1 if request.form.get("is_active", "1") == "1" else 0

            if not name:
                flash("اسم البنك مطلوب.", "danger")
            elif not gl_account_id:
                flash("اختر حساب الأستاذ المرتبط.", "danger")
            else:
                cur.execute(
                    """
                    INSERT INTO banks(
                        name,branch,account_number,iban,currency,
                        opening_balance,current_balance,gl_account_id,is_active,created_by,updated_at
                    )
                    VALUES (?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)
                    """,
                    (
                        name,
                        branch,
                        account_number,
                        iban,
                        currency,
                        _safe_float(opening_balance),
                        _safe_float(opening_balance),
                        gl_account_id,
                        is_active,
                        session.get("username", ""),
                    ),
                )
                bank_id = cur.lastrowid
                log_action(cur, "create", "bank", bank_id, f"bank={name}")
                conn.commit()
                conn.close()
                flash("تم حفظ البنك بنجاح.", "success")
                return redirect(url_for("banks"))

        rows = _fetch_banks(cur)
        accounts = _fetch_accounts(cur)
        conn.close()
        return render_template("banks/list.html", rows=rows, accounts=accounts)

    return banks


def build_bank_edit_view(deps):
    db = deps["db"]
    log_action = deps["log_action"]
    row_snapshot = deps["row_snapshot"]

    def edit_bank(bank_id):
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)
        bank = _bank_record(cur, bank_id)
        if not bank:
            conn.close()
            flash("البنك غير موجود.", "danger")
            return redirect(url_for("banks"))

        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            branch = (request.form.get("branch") or "").strip()
            account_number = (request.form.get("account_number") or "").strip()
            iban = (request.form.get("iban") or "").strip()
            currency = (request.form.get("currency") or "EGP").strip() or "EGP"
            gl_account_id = request.form.get("gl_account_id") or None
            is_active = 1 if request.form.get("is_active", "1") == "1" else 0
            if not name:
                flash("اسم البنك مطلوب.", "danger")
            elif not gl_account_id:
                flash("اختر حساب الأستاذ المرتبط.", "danger")
            else:
                before = row_snapshot(cur, "banks", bank_id)
                cur.execute(
                    """
                    UPDATE banks
                    SET name=?,branch=?,account_number=?,iban=?,currency=?,gl_account_id=?,is_active=?,updated_at=CURRENT_TIMESTAMP
                    WHERE id=?
                    """,
                    (name, branch, account_number, iban, currency, gl_account_id, is_active, bank_id),
                )
                after = row_snapshot(cur, "banks", bank_id)
                log_action(cur, "update", "bank", bank_id, f"bank={name}", before, after)
                conn.commit()
                conn.close()
                flash("تم تحديث بيانات البنك.", "success")
                return redirect(url_for("banks"))

        accounts = _fetch_accounts(cur)
        conn.close()
        return render_template("banks/edit.html", bank=bank, accounts=accounts)

    return edit_bank


def build_bank_toggle_view(deps):
    db = deps["db"]
    log_action = deps["log_action"]

    def toggle_bank(bank_id):
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)
        cur.execute("SELECT name,is_active FROM banks WHERE id=?", (bank_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            flash("البنك غير موجود.", "danger")
            return redirect(url_for("banks"))
        new_status = 0 if row[1] else 1
        cur.execute("UPDATE banks SET is_active=?,updated_at=CURRENT_TIMESTAMP WHERE id=?", (new_status, bank_id))
        log_action(cur, "update", "bank", bank_id, f"toggle active {row[0]} -> {new_status}")
        conn.commit()
        conn.close()
        flash("تم تحديث حالة البنك.", "success")
        return redirect(url_for("banks"))

    return toggle_bank


def build_bank_delete_view(deps):
    db = deps["db"]
    log_action = deps["log_action"]
    row_snapshot = deps["row_snapshot"]

    def delete_bank(bank_id):
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)
        cur.execute("SELECT name FROM banks WHERE id=?", (bank_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            flash("البنك غير موجود.", "danger")
            return redirect(url_for("banks"))
        cur.execute("SELECT COUNT(*) FROM bank_transactions WHERE bank_id=? OR related_bank_id=?", (bank_id, bank_id))
        if cur.fetchone()[0]:
            conn.close()
            flash("لا يمكن حذف بنك له حركات مسجلة. يمكنك تعطيله فقط.", "danger")
            return redirect(url_for("banks"))
        before = row_snapshot(cur, "banks", bank_id)
        cur.execute("DELETE FROM banks WHERE id=?", (bank_id,))
        log_action(cur, "delete", "bank", bank_id, f"delete bank {row[0]}", before, None)
        conn.commit()
        conn.close()
        flash("تم حذف البنك.", "success")
        return redirect(url_for("banks"))

    return delete_bank


def build_bank_transactions_view(deps):
    db = deps["db"]
    next_document_number = deps["next_document_number"]
    ensure_open_period = deps["ensure_open_period"]
    log_action = deps["log_action"]
    mark_journal_source = deps["mark_journal_source"]
    rebuild_ledger = deps["rebuild_ledger"]

    def bank_transactions():
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)

        if request.method == "POST":
            txn_type = (request.form.get("txn_type") or "").strip()
            bank_id = request.form.get("bank_id") or None
            txn_date = (request.form.get("txn_date") or "").strip()
            doc_no = (request.form.get("doc_no") or "").strip() or next_document_number(cur, "bank_transactions")
            amount = _safe_float(request.form.get("amount"))
            description = (request.form.get("description") or "").strip()
            reference_no = (request.form.get("reference_no") or "").strip()
            counterparty_account_id = request.form.get("counterparty_account_id") or None
            notes = (request.form.get("notes") or "").strip()

            bank = _bank_record(cur, bank_id) if bank_id else None
            if txn_type not in {item[0] for item in TRANSACTION_TYPES if item[0] != "bank_to_bank_out"}:
                flash("نوع الحركة غير صحيح.", "danger")
            elif not bank:
                flash("اختر بنكًا صالحًا.", "danger")
            elif not txn_date:
                flash("التاريخ مطلوب.", "danger")
            elif amount <= 0:
                flash("المبلغ يجب أن يكون أكبر من صفر.", "danger")
            else:
                try:
                    ensure_open_period(cur, txn_date)
                    debit_account_id = None
                    credit_account_id = None
                    signed_amount = 0
                    related_bank_id = None
                    if txn_type == "deposit":
                        if not counterparty_account_id:
                            raise ValueError("اختر الحساب المقابل للإيداع.")
                        debit_account_id = bank[8]
                        credit_account_id = int(counterparty_account_id)
                        signed_amount = amount
                    elif txn_type == "withdrawal":
                        if not counterparty_account_id:
                            raise ValueError("اختر الحساب المقابل للسحب.")
                        debit_account_id = int(counterparty_account_id)
                        credit_account_id = bank[8]
                        signed_amount = -amount
                    elif txn_type == "treasury_to_bank":
                        debit_account_id = bank[8]
                        credit_account_id = int(counterparty_account_id or _default_treasury_account_id(cur))
                        signed_amount = amount
                    elif txn_type == "bank_to_treasury":
                        debit_account_id = int(counterparty_account_id or _default_treasury_account_id(cur))
                        credit_account_id = bank[8]
                        signed_amount = -amount
                    elif txn_type == "bank_expense":
                        debit_account_id = int(counterparty_account_id or _system_account_id(cur, "bank_commission"))
                        credit_account_id = bank[8]
                        signed_amount = -amount
                    elif txn_type == "bank_commission":
                        debit_account_id = int(counterparty_account_id or _system_account_id(cur, "bank_commission"))
                        credit_account_id = bank[8]
                        signed_amount = -amount
                    elif txn_type == "bank_interest":
                        debit_account_id = bank[8]
                        credit_account_id = int(counterparty_account_id or _system_account_id(cur, "bank_interest"))
                        signed_amount = amount
                    elif txn_type == "bank_reconciliation":
                        if not counterparty_account_id:
                            counterparty_account_id = _system_account_id(cur, "reconciliation_diff")
                        if amount >= 0:
                            debit_account_id = bank[8]
                            credit_account_id = int(counterparty_account_id)
                            signed_amount = amount
                        else:
                            debit_account_id = int(counterparty_account_id)
                            credit_account_id = bank[8]
                            signed_amount = amount
                    elif txn_type == "opening_balance":
                        debit_account_id = bank[8]
                        credit_account_id = int(counterparty_account_id or _system_account_id(cur, "bank_opening"))
                        signed_amount = amount
                    else:
                        raise ValueError("نوع الحركة غير مدعوم.")

                    journal_id = _post_entry_by_ids(
                        cur,
                        txn_date,
                        description or dict(TRANSACTION_TYPES).get(txn_type, "حركة بنكية"),
                        debit_account_id,
                        credit_account_id,
                        abs(amount),
                        "bank_transaction",
                    )
                    txn_id = _insert_bank_transaction(
                        cur,
                        int(bank_id),
                        txn_type,
                        doc_no,
                        txn_date,
                        description or dict(TRANSACTION_TYPES).get(txn_type, "حركة بنكية"),
                        abs(amount),
                        signed_amount,
                        counterparty_account_id=int(counterparty_account_id) if counterparty_account_id else None,
                        journal_id=journal_id,
                        related_bank_id=related_bank_id,
                        reference_no=reference_no,
                        notes=notes,
                        source_type="bank_transaction",
                    )
                    if journal_id:
                        mark_journal_source(cur, "bank_transaction", txn_id, journal_id)
                    log_action(cur, "create", "bank_transaction", txn_id, f"{txn_type}:{doc_no}")
                    conn.commit()
                    rebuild_ledger()
                    conn.close()
                    flash("تم تسجيل الحركة البنكية.", "success")
                    return redirect(url_for("bank_transactions"))
                except (ValueError, sqlite3.IntegrityError) as exc:
                    conn.rollback()
                    flash(str(exc) if isinstance(exc, ValueError) else "رقم المستند مستخدم بالفعل لنفس نوع الحركة.", "danger")

        banks = _fetch_banks(cur, active_only=True)
        accounts = _fetch_accounts(cur)
        cur.execute(
            """
            SELECT bt.id,b.name,bt.txn_date,bt.doc_no,bt.txn_type,bt.amount,bt.balance_after,bt.status
            FROM bank_transactions bt
            JOIN banks b ON b.id=bt.bank_id
            ORDER BY bt.id DESC
            LIMIT 50
            """
        )
        rows = cur.fetchall()
        conn.close()
        return render_template(
            "banks/transactions.html",
            banks=banks,
            accounts=accounts,
            rows=rows,
            transaction_types=TRANSACTION_TYPES,
        )

    return bank_transactions


def build_bank_transfers_view(deps):
    db = deps["db"]
    next_document_number = deps["next_document_number"]
    ensure_open_period = deps["ensure_open_period"]
    log_action = deps["log_action"]
    mark_journal_source = deps["mark_journal_source"]
    rebuild_ledger = deps["rebuild_ledger"]

    def bank_transfers():
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)
        if request.method == "POST":
            source_bank_id = request.form.get("source_bank_id") or None
            target_bank_id = request.form.get("target_bank_id") or None
            txn_date = (request.form.get("txn_date") or "").strip()
            amount = _safe_float(request.form.get("amount"))
            doc_no = (request.form.get("doc_no") or "").strip() or next_document_number(cur, "bank_transfers")
            reference_no = (request.form.get("reference_no") or "").strip()
            description = (request.form.get("description") or "").strip() or "تحويل بنكي"
            notes = (request.form.get("notes") or "").strip()
            source_bank = _bank_record(cur, source_bank_id) if source_bank_id else None
            target_bank = _bank_record(cur, target_bank_id) if target_bank_id else None
            if not source_bank or not target_bank:
                flash("اختر البنك المحول منه وإليه.", "danger")
            elif source_bank_id == target_bank_id:
                flash("لا يمكن التحويل إلى نفس البنك.", "danger")
            elif not txn_date:
                flash("التاريخ مطلوب.", "danger")
            elif amount <= 0:
                flash("المبلغ يجب أن يكون أكبر من صفر.", "danger")
            else:
                try:
                    ensure_open_period(cur, txn_date)
                    journal_id = _post_entry_by_ids(
                        cur,
                        txn_date,
                        description,
                        target_bank[8],
                        source_bank[8],
                        amount,
                        "bank_transfer",
                    )
                    group_key = f"transfer:{doc_no}"
                    out_txn_id = _insert_bank_transaction(
                        cur,
                        int(source_bank_id),
                        "bank_to_bank_out",
                        doc_no,
                        txn_date,
                        description,
                        amount,
                        -amount,
                        journal_id=journal_id,
                        related_bank_id=int(target_bank_id),
                        reference_no=reference_no,
                        notes=notes,
                        txn_group=group_key,
                        source_type="bank_transfer",
                    )
                    _insert_bank_transaction(
                        cur,
                        int(target_bank_id),
                        "deposit",
                        doc_no,
                        txn_date,
                        description,
                        amount,
                        amount,
                        journal_id=journal_id,
                        related_bank_id=int(source_bank_id),
                        reference_no=reference_no,
                        notes=notes,
                        txn_group=group_key,
                        source_type="bank_transfer",
                    )
                    if journal_id:
                        mark_journal_source(cur, "bank_transfer", out_txn_id, journal_id)
                    log_action(cur, "create", "bank_transfer", out_txn_id, f"transfer:{doc_no}")
                    conn.commit()
                    rebuild_ledger()
                    conn.close()
                    flash("تم تنفيذ التحويل البنكي.", "success")
                    return redirect(url_for("bank_transfers"))
                except (ValueError, sqlite3.IntegrityError) as exc:
                    conn.rollback()
                    flash(str(exc) if isinstance(exc, ValueError) else "رقم المستند مستخدم بالفعل لنفس نوع الحركة.", "danger")

        banks = _fetch_banks(cur, active_only=True)
        cur.execute(
            """
            SELECT bt.id,b1.name,b2.name,bt.txn_date,bt.doc_no,bt.amount,bt.reference_no
            FROM bank_transactions bt
            JOIN banks b1 ON b1.id=bt.bank_id
            LEFT JOIN banks b2 ON b2.id=bt.related_bank_id
            WHERE bt.txn_type='bank_to_bank_out'
            ORDER BY bt.id DESC
            LIMIT 50
            """
        )
        rows = cur.fetchall()
        conn.close()
        return render_template("banks/transfers.html", banks=banks, rows=rows)

    return bank_transfers


def build_bank_statement_view(deps):
    db = deps["db"]

    def bank_statement():
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)
        bank_id = (request.args.get("bank_id") or "").strip()
        from_date = (request.args.get("from_date") or "").strip()
        to_date = (request.args.get("to_date") or "").strip()
        txn_type = (request.args.get("txn_type") or "").strip()
        banks = _fetch_banks(cur, active_only=True)
        rows = []
        opening_balance = None
        bank = None
        if bank_id:
            bank = _bank_record(cur, bank_id)
            if bank:
                opening_balance = _statement_opening_balance(cur, int(bank_id), from_date or None)
                rows = _bank_statement_rows(cur, int(bank_id), from_date, to_date, txn_type)
        export_rows = [
            (
                row[1],
                row[2],
                dict(TRANSACTION_TYPES).get(row[3], row[3]),
                row[4],
                row[5],
                row[6],
                row[7],
                row[8],
                row[9],
            )
            for row in rows
        ]
        if request.args.get("format") == "excel" and bank:
            conn.close()
            return _statement_excel(
                f"bank_statement_{bank_id}_{from_date or 'all'}_{to_date or 'all'}.xlsx",
                f"كشف حساب البنك - {bank[1]}",
                ["التاريخ", "رقم المستند", "نوع الحركة", "الوصف", "المرجع", "المبلغ", "الأثر", "رصيد قبل", "رصيد بعد"],
                export_rows,
            )
        conn.close()
        return render_template(
            "banks/statement.html",
            banks=banks,
            rows=rows,
            bank=bank,
            opening_balance=opening_balance,
            from_date=from_date,
            to_date=to_date,
            txn_type=txn_type,
            transaction_types=TRANSACTION_TYPES,
        )

    return bank_statement


def build_bank_reconciliation_view(deps):
    db = deps["db"]
    next_document_number = deps["next_document_number"]
    mark_journal_source = deps["mark_journal_source"]
    rebuild_ledger = deps["rebuild_ledger"]
    log_action = deps["log_action"]

    def bank_reconciliation():
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)

        if request.method == "POST":
            bank_id = request.form.get("bank_id") or None
            statement_date = (request.form.get("statement_date") or "").strip()
            from_date = (request.form.get("from_date") or "").strip()
            to_date = (request.form.get("to_date") or "").strip()
            statement_balance = _safe_float(request.form.get("statement_balance"))
            notes = (request.form.get("notes") or "").strip()
            lines_text = request.form.get("lines_text", "")
            post_difference = request.form.get("post_difference") == "1"
            difference_account_id = request.form.get("difference_account_id") or _system_account_id(cur, "reconciliation_diff")
            bank = _bank_record(cur, bank_id) if bank_id else None
            if not bank:
                flash("اختر بنكًا صالحًا.", "danger")
            elif not statement_date:
                flash("تاريخ الكشف مطلوب.", "danger")
            else:
                doc_no = (request.form.get("doc_no") or "").strip() or next_document_number(cur, "bank_reconciliations")
                opening_balance = _statement_opening_balance(cur, int(bank_id), from_date or None)
                rows = _bank_statement_rows(cur, int(bank_id), from_date, to_date, "")
                system_balance = opening_balance if not rows else rows[-1][9]
                difference_amount = round(statement_balance - float(system_balance or 0), 2)
                journal_id = None
                try:
                    if post_difference and abs(difference_amount) > 0.0001:
                        if difference_amount > 0:
                            journal_id = _post_entry_by_ids(
                                cur,
                                statement_date,
                                f"تسوية بنكية {doc_no}",
                                bank[8],
                                int(difference_account_id),
                                difference_amount,
                                "bank_reconciliation",
                            )
                            _insert_bank_transaction(
                                cur,
                                int(bank_id),
                                "bank_reconciliation",
                                doc_no,
                                statement_date,
                                f"تسوية بنكية {doc_no}",
                                difference_amount,
                                difference_amount,
                                counterparty_account_id=int(difference_account_id),
                                journal_id=journal_id,
                                notes=notes,
                                source_type="bank_reconciliation",
                            )
                        else:
                            journal_id = _post_entry_by_ids(
                                cur,
                                statement_date,
                                f"تسوية بنكية {doc_no}",
                                int(difference_account_id),
                                bank[8],
                                abs(difference_amount),
                                "bank_reconciliation",
                            )
                            _insert_bank_transaction(
                                cur,
                                int(bank_id),
                                "bank_reconciliation",
                                doc_no,
                                statement_date,
                                f"تسوية بنكية {doc_no}",
                                abs(difference_amount),
                                difference_amount,
                                counterparty_account_id=int(difference_account_id),
                                journal_id=journal_id,
                                notes=notes,
                                source_type="bank_reconciliation",
                            )
                    cur.execute(
                        """
                        INSERT INTO bank_reconciliations(
                            bank_id,doc_no,statement_date,from_date,to_date,opening_balance,
                            system_balance,statement_balance,difference_amount,status,notes,journal_id,created_by
                        )
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            bank_id,
                            doc_no,
                            statement_date,
                            from_date,
                            to_date,
                            opening_balance,
                            system_balance,
                            statement_balance,
                            difference_amount,
                            "posted" if post_difference else "draft",
                            notes,
                            journal_id,
                            session.get("username", ""),
                        ),
                    )
                    reconciliation_id = cur.lastrowid
                    if journal_id:
                        mark_journal_source(cur, "bank_reconciliation", reconciliation_id, journal_id)
                    for line_date, amount, reference_no, description in _parse_reconciliation_lines(lines_text):
                        match = _match_reconciliation_line(cur, int(bank_id), line_date, amount, reference_no)
                        matched_transaction_id = match[0] if match else None
                        match_status = "matched" if match else "unmatched"
                        diff = 0 if match else amount
                        cur.execute(
                            """
                            INSERT INTO bank_reconciliation_lines(
                                reconciliation_id,line_date,description,reference_no,amount,
                                matched_transaction_id,match_status,difference_amount,notes
                            )
                            VALUES (?,?,?,?,?,?,?,?,?)
                            """,
                            (
                                reconciliation_id,
                                line_date,
                                description,
                                reference_no,
                                amount,
                                matched_transaction_id,
                                match_status,
                                diff,
                                "",
                            ),
                        )
                    log_action(cur, "create", "bank_reconciliation", reconciliation_id, f"reconciliation:{doc_no}")
                    conn.commit()
                    rebuild_ledger()
                    conn.close()
                    flash("تم تسجيل التسوية البنكية.", "success")
                    return redirect(url_for("bank_reconciliation_detail", reconciliation_id=reconciliation_id))
                except sqlite3.IntegrityError:
                    conn.rollback()
                    flash("رقم مستند التسوية مستخدم بالفعل.", "danger")

        banks = _fetch_banks(cur, active_only=True)
        accounts = _fetch_accounts(cur)
        cur.execute(
            """
            SELECT br.id,b.name,br.doc_no,br.statement_date,br.system_balance,br.statement_balance,
                   br.difference_amount,br.status
            FROM bank_reconciliations br
            JOIN banks b ON b.id=br.bank_id
            ORDER BY br.id DESC
            LIMIT 50
            """
        )
        rows = cur.fetchall()
        conn.close()
        return render_template("banks/reconciliation.html", banks=banks, accounts=accounts, rows=rows)

    return bank_reconciliation


def build_bank_reconciliation_detail_view(deps):
    db = deps["db"]

    def bank_reconciliation_detail(reconciliation_id):
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)
        cur.execute(
            """
            SELECT br.id,br.doc_no,br.statement_date,br.from_date,br.to_date,br.opening_balance,
                   br.system_balance,br.statement_balance,br.difference_amount,br.status,br.notes,b.name
            FROM bank_reconciliations br
            JOIN banks b ON b.id=br.bank_id
            WHERE br.id=?
            """,
            (reconciliation_id,),
        )
        reconciliation = cur.fetchone()
        if not reconciliation:
            conn.close()
            flash("التسوية البنكية غير موجودة.", "danger")
            return redirect(url_for("bank_reconciliation"))
        cur.execute(
            """
            SELECT brl.line_date,brl.description,brl.reference_no,brl.amount,brl.match_status,
                   bt.doc_no,bt.description,brl.difference_amount
            FROM bank_reconciliation_lines brl
            LEFT JOIN bank_transactions bt ON bt.id=brl.matched_transaction_id
            WHERE brl.reconciliation_id=?
            ORDER BY brl.id
            """,
            (reconciliation_id,),
        )
        lines = cur.fetchall()
        conn.close()
        return render_template("banks/reconciliation_detail.html", reconciliation=reconciliation, lines=lines)

    return bank_reconciliation_detail


def build_receivable_checks_view(deps):
    db = deps["db"]
    next_document_number = deps["next_document_number"]
    mark_journal_source = deps["mark_journal_source"]
    rebuild_ledger = deps["rebuild_ledger"]
    log_action = deps["log_action"]

    def receivable_checks():
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)
        if request.method == "POST":
            doc_no = (request.form.get("doc_no") or "").strip() or next_document_number(cur, "receivable_checks")
            check_number = (request.form.get("check_number") or "").strip()
            bank_id = request.form.get("bank_id") or None
            customer_id = request.form.get("customer_id") or None
            due_date = (request.form.get("due_date") or "").strip()
            amount = _safe_float(request.form.get("amount"))
            notes = (request.form.get("notes") or "").strip()
            bank = _bank_record(cur, bank_id) if bank_id else None
            if not check_number or not bank or not customer_id or not due_date or amount <= 0:
                flash("أكمل بيانات الشيك الوارد بشكل صحيح.", "danger")
            else:
                try:
                    journal_id = _post_entry_by_ids(
                        cur,
                        due_date,
                        f"استلام شيك وارد {check_number}",
                        _system_account_id(cur, "receivable_checks"),
                        _account_id_by_code(cur, "1300"),
                        amount,
                        "receivable_check",
                    )
                    cur.execute(
                        """
                        INSERT INTO receivable_checks(
                            doc_no,check_number,bank_id,customer_id,due_date,amount,status,notes,received_journal_id,created_by
                        )
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            doc_no,
                            check_number,
                            bank_id,
                            customer_id,
                            due_date,
                            amount,
                            "portfolio",
                            notes,
                            journal_id,
                            session.get("username", ""),
                        ),
                    )
                    check_id = cur.lastrowid
                    if journal_id:
                        mark_journal_source(cur, "receivable_check", check_id, journal_id)
                    log_action(cur, "create", "receivable_check", check_id, f"check={check_number}")
                    conn.commit()
                    rebuild_ledger()
                    conn.close()
                    flash("تم تسجيل الشيك الوارد.", "success")
                    return redirect(url_for("receivable_checks"))
                except sqlite3.IntegrityError:
                    conn.rollback()
                    flash("رقم الشيك مستخدم بالفعل لنفس البنك.", "danger")

        banks = _fetch_banks(cur, active_only=True)
        customers = _fetch_customers(cur)
        rows, due, overdue, bounced, status_map = _build_check_summary(cur, "receivable_checks", customer_side=True)
        conn.close()
        return render_template(
            "banks/receivable_checks.html",
            banks=banks,
            customers=customers,
            rows=rows,
            due_rows=due,
            overdue_rows=overdue,
            bounced_rows=bounced,
            status_map=status_map,
        )

    return receivable_checks


def build_receivable_check_action_view(deps):
    db = deps["db"]
    next_document_number = deps["next_document_number"]
    mark_journal_source = deps["mark_journal_source"]
    rebuild_ledger = deps["rebuild_ledger"]

    def receivable_check_action(check_id):
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)
        cur.execute(
            """
            SELECT id,doc_no,check_number,bank_id,customer_id,due_date,amount,status,received_journal_id,deposit_journal_id,bank_transaction_id
            FROM receivable_checks
            WHERE id=?
            """,
            (check_id,),
        )
        check = cur.fetchone()
        if not check:
            conn.close()
            flash("الشيك غير موجود.", "danger")
            return redirect(url_for("receivable_checks"))
        action = (request.form.get("action") or "").strip()
        notes = (request.form.get("notes") or "").strip()
        bank = _bank_record(cur, check[3])
        try:
            if action == "deposit" and check[7] == "portfolio":
                doc_no = (request.form.get("doc_no") or "").strip() or next_document_number(cur, "bank_transactions")
                journal_id = _post_entry_by_ids(
                    cur,
                    date.today().isoformat(),
                    f"إيداع شيك وارد {check[2]}",
                    bank[8],
                    _system_account_id(cur, "receivable_checks"),
                    check[6],
                    "receivable_check_deposit",
                )
                txn_id = _insert_bank_transaction(
                    cur,
                    check[3],
                    "deposit",
                    doc_no,
                    date.today().isoformat(),
                    f"إيداع شيك وارد {check[2]}",
                    check[6],
                    check[6],
                    counterparty_account_id=_system_account_id(cur, "receivable_checks"),
                    journal_id=journal_id,
                    reference_no=check[2],
                    notes=notes,
                    source_type="receivable_check",
                    source_id=check_id,
                )
                cur.execute(
                    "UPDATE receivable_checks SET status='deposited',deposit_journal_id=?,bank_transaction_id=?,notes=COALESCE(notes,'') || ? WHERE id=?",
                    (journal_id, txn_id, ("\n" + notes) if notes else "", check_id),
                )
                if journal_id:
                    mark_journal_source(cur, "receivable_check", check_id, journal_id)
            elif action == "collect" and check[7] in ("portfolio", "deposited"):
                if check[7] == "portfolio":
                    doc_no = (request.form.get("doc_no") or "").strip() or next_document_number(cur, "bank_transactions")
                    journal_id = _post_entry_by_ids(
                        cur,
                        date.today().isoformat(),
                        f"تحصيل شيك وارد {check[2]}",
                        bank[8],
                        _system_account_id(cur, "receivable_checks"),
                        check[6],
                        "receivable_check_collect",
                    )
                    txn_id = _insert_bank_transaction(
                        cur,
                        check[3],
                        "deposit",
                        doc_no,
                        date.today().isoformat(),
                        f"تحصيل شيك وارد {check[2]}",
                        check[6],
                        check[6],
                        counterparty_account_id=_system_account_id(cur, "receivable_checks"),
                        journal_id=journal_id,
                        reference_no=check[2],
                        notes=notes,
                        source_type="receivable_check",
                        source_id=check_id,
                    )
                    cur.execute(
                        "UPDATE receivable_checks SET status='collected',deposit_journal_id=?,bank_transaction_id=? WHERE id=?",
                        (journal_id, txn_id, check_id),
                    )
                    if journal_id:
                        mark_journal_source(cur, "receivable_check", check_id, journal_id)
                else:
                    cur.execute("UPDATE receivable_checks SET status='collected' WHERE id=?", (check_id,))
            elif action == "bounce" and check[7] in ("portfolio", "deposited", "collected"):
                journal_ids = []
                if check[7] in ("deposited", "collected"):
                    reverse_doc_no = (request.form.get("doc_no") or "").strip() or next_document_number(cur, "bank_transactions")
                    deposit_reversal_journal_id = _post_entry_by_ids(
                        cur,
                        date.today().isoformat(),
                        f"عكس إيداع شيك مرتجع {check[2]}",
                        _system_account_id(cur, "receivable_checks"),
                        bank[8],
                        check[6],
                        "receivable_check_bounce",
                    )
                    journal_ids.append(deposit_reversal_journal_id)
                    reversal_txn_id = _insert_bank_transaction(
                        cur,
                        check[3],
                        "withdrawal",
                        reverse_doc_no,
                        date.today().isoformat(),
                        f"عكس إيداع شيك مرتجع {check[2]}",
                        check[6],
                        -check[6],
                        counterparty_account_id=_system_account_id(cur, "receivable_checks"),
                        journal_id=deposit_reversal_journal_id,
                        reference_no=check[2],
                        notes=notes,
                        source_type="receivable_check",
                        source_id=check_id,
                    )
                journal_ids.append(
                    _post_entry_by_ids(
                        cur,
                        date.today().isoformat(),
                        f"إثبات رفض شيك {check[2]}",
                        _account_id_by_code(cur, "1300"),
                        _system_account_id(cur, "receivable_checks"),
                        check[6],
                        "receivable_check_bounce",
                    )
                )
                cur.execute(
                    "UPDATE receivable_checks SET status='bounced',bounce_journal_id=?,bank_transaction_id=COALESCE(bank_transaction_id, ?) WHERE id=?",
                    (journal_ids[-1], locals().get("reversal_txn_id"), check_id),
                )
                for journal_id in journal_ids:
                    if journal_id:
                        mark_journal_source(cur, "receivable_check", check_id, journal_id)
            else:
                flash("لا يمكن تنفيذ الإجراء المطلوب على هذه الحالة.", "danger")
                conn.close()
                return redirect(url_for("receivable_checks"))
            conn.commit()
            rebuild_ledger()
            conn.close()
            flash("تم تحديث حالة الشيك الوارد.", "success")
            return redirect(url_for("receivable_checks"))
        except sqlite3.IntegrityError:
            conn.rollback()
            conn.close()
            flash("رقم المستند مستخدم بالفعل.", "danger")
            return redirect(url_for("receivable_checks"))

    return receivable_check_action


def build_payable_checks_view(deps):
    db = deps["db"]
    next_document_number = deps["next_document_number"]
    mark_journal_source = deps["mark_journal_source"]
    rebuild_ledger = deps["rebuild_ledger"]
    log_action = deps["log_action"]

    def payable_checks():
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)
        if request.method == "POST":
            doc_no = (request.form.get("doc_no") or "").strip() or next_document_number(cur, "payable_checks")
            check_number = (request.form.get("check_number") or "").strip()
            bank_id = request.form.get("bank_id") or None
            supplier_id = request.form.get("supplier_id") or None
            due_date = (request.form.get("due_date") or "").strip()
            amount = _safe_float(request.form.get("amount"))
            notes = (request.form.get("notes") or "").strip()
            bank = _bank_record(cur, bank_id) if bank_id else None
            if not check_number or not bank or not supplier_id or not due_date or amount <= 0:
                flash("أكمل بيانات الشيك الصادر بشكل صحيح.", "danger")
            else:
                try:
                    journal_id = _post_entry_by_ids(
                        cur,
                        due_date,
                        f"إصدار شيك {check_number}",
                        _account_id_by_code(cur, "2100"),
                        _system_account_id(cur, "payable_checks"),
                        amount,
                        "payable_check",
                    )
                    cur.execute(
                        """
                        INSERT INTO payable_checks(
                            doc_no,check_number,bank_id,supplier_id,due_date,amount,status,notes,issue_journal_id,created_by
                        )
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            doc_no,
                            check_number,
                            bank_id,
                            supplier_id,
                            due_date,
                            amount,
                            "issued",
                            notes,
                            journal_id,
                            session.get("username", ""),
                        ),
                    )
                    check_id = cur.lastrowid
                    if journal_id:
                        mark_journal_source(cur, "payable_check", check_id, journal_id)
                    log_action(cur, "create", "payable_check", check_id, f"check={check_number}")
                    conn.commit()
                    rebuild_ledger()
                    conn.close()
                    flash("تم تسجيل الشيك الصادر.", "success")
                    return redirect(url_for("payable_checks"))
                except sqlite3.IntegrityError:
                    conn.rollback()
                    flash("رقم الشيك مستخدم بالفعل لنفس البنك.", "danger")

        banks = _fetch_banks(cur, active_only=True)
        suppliers = _fetch_suppliers(cur)
        rows, due, overdue, cancelled, status_map = _build_check_summary(cur, "payable_checks", customer_side=False)
        conn.close()
        return render_template(
            "banks/payable_checks.html",
            banks=banks,
            suppliers=suppliers,
            rows=rows,
            due_rows=due,
            overdue_rows=overdue,
            cancelled_rows=cancelled,
            status_map=status_map,
        )

    return payable_checks


def build_payable_check_action_view(deps):
    db = deps["db"]
    next_document_number = deps["next_document_number"]
    mark_journal_source = deps["mark_journal_source"]
    rebuild_ledger = deps["rebuild_ledger"]

    def payable_check_action(check_id):
        conn = db()
        cur = conn.cursor()
        _ensure_context(cur)
        cur.execute(
            """
            SELECT id,doc_no,check_number,bank_id,supplier_id,due_date,amount,status,issue_journal_id,cash_journal_id,bank_transaction_id
            FROM payable_checks
            WHERE id=?
            """,
            (check_id,),
        )
        check = cur.fetchone()
        if not check:
            conn.close()
            flash("الشيك غير موجود.", "danger")
            return redirect(url_for("payable_checks"))
        action = (request.form.get("action") or "").strip()
        notes = (request.form.get("notes") or "").strip()
        bank = _bank_record(cur, check[3])
        try:
            if action == "deliver" and check[7] == "issued":
                cur.execute("UPDATE payable_checks SET status='delivered' WHERE id=?", (check_id,))
            elif action == "cash" and check[7] in ("issued", "delivered"):
                doc_no = (request.form.get("doc_no") or "").strip() or next_document_number(cur, "bank_transactions")
                journal_id = _post_entry_by_ids(
                    cur,
                    date.today().isoformat(),
                    f"صرف شيك {check[2]}",
                    _system_account_id(cur, "payable_checks"),
                    bank[8],
                    check[6],
                    "payable_check_cash",
                )
                txn_id = _insert_bank_transaction(
                    cur,
                    check[3],
                    "withdrawal",
                    doc_no,
                    date.today().isoformat(),
                    f"صرف شيك {check[2]}",
                    check[6],
                    -check[6],
                    counterparty_account_id=_system_account_id(cur, "payable_checks"),
                    journal_id=journal_id,
                    reference_no=check[2],
                    notes=notes,
                    source_type="payable_check",
                    source_id=check_id,
                )
                cur.execute(
                    "UPDATE payable_checks SET status='cashed',cash_journal_id=?,bank_transaction_id=? WHERE id=?",
                    (journal_id, txn_id, check_id),
                )
                if journal_id:
                    mark_journal_source(cur, "payable_check", check_id, journal_id)
            elif action == "cancel" and check[7] in ("issued", "delivered"):
                journal_id = _post_entry_by_ids(
                    cur,
                    date.today().isoformat(),
                    f"إلغاء شيك {check[2]}",
                    _system_account_id(cur, "payable_checks"),
                    _account_id_by_code(cur, "2100"),
                    check[6],
                    "payable_check_cancel",
                )
                cur.execute(
                    "UPDATE payable_checks SET status='cancelled',cancel_journal_id=? WHERE id=?",
                    (journal_id, check_id),
                )
                if journal_id:
                    mark_journal_source(cur, "payable_check", check_id, journal_id)
            else:
                flash("لا يمكن تنفيذ الإجراء المطلوب على هذه الحالة.", "danger")
                conn.close()
                return redirect(url_for("payable_checks"))
            conn.commit()
            rebuild_ledger()
            conn.close()
            flash("تم تحديث حالة الشيك الصادر.", "success")
            return redirect(url_for("payable_checks"))
        except sqlite3.IntegrityError:
            conn.rollback()
            conn.close()
            flash("رقم المستند مستخدم بالفعل.", "danger")
            return redirect(url_for("payable_checks"))

    return payable_check_action
