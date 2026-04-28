import openpyxl
from flask import flash, redirect, render_template, request, url_for


def _text(value):
    if value is None:
        return ""
    return str(value).strip()


def _number(value, default=0):
    try:
        if value in (None, ""):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _sheet_headers(sheet):
    headers = []
    for cell in sheet[1]:
        headers.append(_text(cell.value).lower())
    return headers


def _row_to_dict(headers, row):
    data = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        data[header] = row[index] if index < len(row) else None
    return data


def _get(data, *names, default=""):
    for name in names:
        key = name.lower()
        if key in data:
            return data.get(key)
    return default


def import_customers_sheet(cur, sheet):
    headers = _sheet_headers(sheet)
    added = 0
    skipped = 0
    errors = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        data = _row_to_dict(headers, row)

        name = _text(_get(data, "name", "customer_name", "اسم العميل"))
        phone = _text(_get(data, "phone", "mobile", "telephone", "الهاتف"))
        address = _text(_get(data, "address", "العنوان"))
        tax_registration_number = _text(_get(data, "tax_registration_number", "tax_number", "الرقم الضريبي"))
        tax_card_number = _text(_get(data, "tax_card_number", "البطاقة الضريبية"))
        commercial_register = _text(_get(data, "commercial_register", "السجل التجاري"))
        contact_person = _text(_get(data, "contact_person", "مسؤول التواصل"))
        email = _text(_get(data, "email", "البريد الإلكتروني"))
        withholding_status = _text(_get(data, "withholding_status", "حالة الخصم والإضافة")) or "non_subject"

        if not name:
            skipped += 1
            continue

        if withholding_status not in ("subject", "non_subject"):
            errors.append(f"Customers row {row_number}: withholding_status must be subject or non_subject")
            skipped += 1
            continue

        cur.execute("SELECT id FROM customers WHERE name=?", (name,))
        if cur.fetchone():
            skipped += 1
            continue

        cur.execute(
            """
            INSERT INTO customers(
                name, phone, address, tax_registration_number, tax_card_number,
                commercial_register, contact_person, email, withholding_status
            )
            VALUES (?,?,?,?,?,?,?,?,?)
            """,
            (
                name,
                phone,
                address,
                tax_registration_number,
                tax_card_number,
                commercial_register,
                contact_person,
                email,
                withholding_status,
            ),
        )
        added += 1

    return added, skipped, errors


def import_suppliers_sheet(cur, sheet):
    headers = _sheet_headers(sheet)
    added = 0
    skipped = 0
    errors = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        data = _row_to_dict(headers, row)

        name = _text(_get(data, "name", "supplier_name", "اسم المورد"))
        phone = _text(_get(data, "phone", "mobile", "telephone", "الهاتف"))
        address = _text(_get(data, "address", "العنوان"))
        tax_registration_number = _text(_get(data, "tax_registration_number", "tax_number", "الرقم الضريبي"))
        tax_card_number = _text(_get(data, "tax_card_number", "البطاقة الضريبية"))
        commercial_register = _text(_get(data, "commercial_register", "السجل التجاري"))
        contact_person = _text(_get(data, "contact_person", "مسؤول التواصل"))
        email = _text(_get(data, "email", "البريد الإلكتروني"))
        withholding_status = _text(_get(data, "withholding_status", "حالة الخصم والإضافة")) or "exempt"

        if not name:
            skipped += 1
            continue

        if withholding_status not in ("taxable", "exempt"):
            errors.append(f"Suppliers row {row_number}: withholding_status must be taxable or exempt")
            skipped += 1
            continue

        cur.execute("SELECT id FROM suppliers WHERE name=?", (name,))
        if cur.fetchone():
            skipped += 1
            continue

        cur.execute(
            """
            INSERT INTO suppliers(
                name, phone, address, tax_registration_number, tax_card_number,
                commercial_register, contact_person, email, withholding_status
            )
            VALUES (?,?,?,?,?,?,?,?,?)
            """,
            (
                name,
                phone,
                address,
                tax_registration_number,
                tax_card_number,
                commercial_register,
                contact_person,
                email,
                withholding_status,
            ),
        )
        added += 1

    return added, skipped, errors


def import_products_sheet(cur, sheet):
    headers = _sheet_headers(sheet)
    added = 0
    skipped = 0
    errors = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        data = _row_to_dict(headers, row)

        code = _text(_get(data, "code", "product_code", "كود الصنف"))
        name = _text(_get(data, "name", "product_name", "اسم الصنف"))
        unit = _text(_get(data, "unit", "وحدة", "الوحدة")) or "وحدة"
        purchase_price = _number(_get(data, "purchase_price", "cost", "سعر الشراء"), 0)
        sale_price = _number(_get(data, "sale_price", "price", "سعر البيع"), 0)
        stock_quantity = _number(_get(data, "stock_quantity", "quantity", "الرصيد"), 0)

        if not name:
            skipped += 1
            continue

        if code:
            cur.execute("SELECT id FROM products WHERE code=?", (code,))
            if cur.fetchone():
                skipped += 1
                continue
        else:
            cur.execute("SELECT id FROM products WHERE name=?", (name,))
            if cur.fetchone():
                skipped += 1
                continue

        cur.execute(
            """
            INSERT INTO products(code, name, unit, purchase_price, sale_price, stock_quantity)
            VALUES (?,?,?,?,?,?)
            """,
            (
                code or None,
                name,
                unit,
                purchase_price,
                sale_price,
                stock_quantity,
            ),
        )
        added += 1

    return added, skipped, errors


def build_import_full_data_view(deps):
    db = deps["db"]

    def import_full_data():
        if request.method == "POST":
            file = request.files.get("file")

            if not file:
                flash("من فضلك اختر ملف Excel.", "danger")
                return redirect(url_for("import_full_data"))

            filename = (file.filename or "").lower()
            if not filename.endswith(".xlsx"):
                flash("يجب رفع ملف Excel بصيغة .xlsx فقط.", "danger")
                return redirect(url_for("import_full_data"))

            conn = None
            try:
                wb = openpyxl.load_workbook(file, data_only=True)

                conn = db()
                cur = conn.cursor()

                results = {
                    "customers_added": 0,
                    "customers_skipped": 0,
                    "suppliers_added": 0,
                    "suppliers_skipped": 0,
                    "products_added": 0,
                    "products_skipped": 0,
                }
                all_errors = []

                if "Customers" in wb.sheetnames:
                    added, skipped, errors = import_customers_sheet(cur, wb["Customers"])
                    results["customers_added"] = added
                    results["customers_skipped"] = skipped
                    all_errors.extend(errors)

                if "Suppliers" in wb.sheetnames:
                    added, skipped, errors = import_suppliers_sheet(cur, wb["Suppliers"])
                    results["suppliers_added"] = added
                    results["suppliers_skipped"] = skipped
                    all_errors.extend(errors)

                if "Products" in wb.sheetnames:
                    added, skipped, errors = import_products_sheet(cur, wb["Products"])
                    results["products_added"] = added
                    results["products_skipped"] = skipped
                    all_errors.extend(errors)

                conn.commit()
                conn.close()
                conn = None

                message = (
                    f"تم الاستيراد بنجاح: "
                    f"عملاء مضافين {results['customers_added']} / متجاهلين {results['customers_skipped']} - "
                    f"موردين مضافين {results['suppliers_added']} / متجاهلين {results['suppliers_skipped']} - "
                    f"أصناف مضافة {results['products_added']} / متجاهلة {results['products_skipped']}"
                )

                if all_errors:
                    message += " | ملاحظات: " + " - ".join(all_errors[:5])

                flash(message, "success" if not all_errors else "warning")

            except Exception as exc:
                if conn is not None:
                    conn.close()
                flash(f"حدث خطأ أثناء الاستيراد: {exc}", "danger")

            return redirect(url_for("import_full_data"))

        return render_template("import_full_data.html")

    return import_full_data
